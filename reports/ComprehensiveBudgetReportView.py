# reports/views.py
import logging
from decimal import Decimal
import jdatetime

from django.db.models import Sum, Q, Count, OuterRef, Subquery, DecimalField, Prefetch
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.template import context
from django.template.loader import render_to_string
from django.template.smartif import infix
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, View
from django.contrib import messages
from rest_framework.views import APIView

# مدل‌های برنامه
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from core.PermissionBase import PermissionBaseView
from core.models import Organization, Project
from tankhah.models import Tankhah, Factor

# توابع محاسباتی
from budgets.budget_calculations import get_tankhah_used_budget

from django.db.models import Sum, Count, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views.generic import ListView
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils import formats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import jdatetime
import logging
from decimal import Decimal
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from django.db.models.expressions import OuterRef, Subquery

# reports/views.py

import logging
from decimal import Decimal
from django.db.models import Sum, Count, Q, OuterRef, Subquery, DecimalField
from django.http import JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import ListView
import jdatetime

# مدل‌های خود را import کنید
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from core.models import Organization, Project
from tankhah.models import Tankhah, Factor

logger = logging.getLogger(__name__)

# ===================================================================
# 1. ویوی اصلی گزارش (این ویو را دست نزنید، به نظر درست می‌آید)
# ===================================================================

class ComprehensiveBudgetReportView(PermissionBaseView, ListView):
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Starting get_queryset...")
        try:
            total_allocated_subquery = BudgetAllocation.objects.filter(
                budget_period_id=OuterRef('pk'), is_active=True
            ).values('budget_period_id').annotate(total=Sum('allocated_amount')).values('total')
            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='CONSUMPTION'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')
            returned_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='RETURN'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')
            queryset = BudgetPeriod.objects.filter(
                is_active=True, is_completed=False
            ).select_related('organization').annotate(
                period_total_allocated=Coalesce(Subquery(total_allocated_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('-start_date', 'name')
            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"Applying search filter: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)
            count = queryset.count()
            logger.info(f"Found {count} BudgetPeriods.")
            if count == 0:
                logger.warning("No active BudgetPeriods found.")
            return queryset
        except Exception as e:
            logger.error(f"Error in get_queryset: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در بارگذاری دوره‌های بودجه رخ داد."))
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        logger.info(f"[{self.__class__.__name__}] - Starting get_context_data...")
        context = super().get_context_data(**kwargs)
        periods = context.get('object_list', [])
        processed_data = []
        for period in periods:
            net_consumed = period.period_total_consumed - period.period_total_returned
            remaining = period.total_amount - net_consumed
            utilization = (net_consumed / period.total_amount * 100) if period.total_amount else 0
            org_summaries = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                org_total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                org_consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                org_returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='RETURN'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('organization__name')

            org_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'code': org['organization__code'],
                    'total_allocated': org['org_total_allocated'] or 0,
                    'total_allocated_formatted': f"{org['org_total_allocated'] or 0:,.0f}",
                    'net_consumed': (org['org_consumed'] - org['org_returned']) or 0,
                    'net_consumed_formatted': f"{(org['org_consumed'] - org['org_returned']) or 0:,.0f}",
                    'remaining': (org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0,
                    'remaining_formatted': f"{(org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0:,.0f}",
                    'num_budget_items': org['num_budget_items'],
                    'utilization_percentage': (
                        (org['org_consumed'] - org['org_returned']) / org['org_total_allocated'] * 100
                    ) if org['org_total_allocated'] else 0,
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']})
                } for org in org_summaries
            ]
            logger.info(f'org_summaries IS : {org_summaries}')
            processed_data.append({
                'period': period,
                'summary': {
                    'total_budget': period.total_amount or 0,
                    'total_allocated': period.period_total_allocated or 0,
                    'net_consumed': net_consumed or 0,
                    'remaining': remaining or 0,
                    'utilization_percentage': utilization,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime(
                        '%Y/%m/%d') if period.start_date else '-',
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime(
                        '%Y/%m/%d') if period.end_date else '-',
                },
                'organization_summaries': org_data
            })
        context[self.context_object_name] = processed_data
        context['report_main_title'] = _("گزارش جامع بودجه")
        context['current_search_period'] = self.request.GET.get('search_period', '')
        if not processed_data:
            messages.info(self.request, _("هیچ دوره بودجه فعالی یافت نشد."))
        return context

    def generate_excel_report(self, report_data):
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl is not installed.")
            messages.error(self.request, _("لطفاً openpyxl را نصب کنید."))
            return None
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="budget_report.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True
        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل (ریال)"),
            _("تخصیص (ریال)"), _("مصرف (ریال)"), _("مانده (ریال)"), _("% مصرف"),
            _("سازمان تخصیص"), _("کد"), _("تخصیص سازمان (ریال)"), _("مصرف سازمان (ریال)"),
            _("مانده سازمان (ریال)"), _("% مصرف سازمان")
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col)].width = 20
        row = 2
        for period_data in report_data:
            period = period_data['period']
            summary = period_data['summary']
            start_row = row
            if not period_data['organization_summaries']:
                sheet.cell(row=row, column=1, value=period.name)
                sheet.cell(row=row, column=2, value=period.organization.name)
                sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                row += 1
            else:
                for org in period_data['organization_summaries']:
                    sheet.cell(row=row, column=1, value=period.name)
                    sheet.cell(row=row, column=2, value=period.organization.name)
                    sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                    sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                    sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                    sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                    sheet.cell(row=row, column=10, value=org['name'])
                    sheet.cell(row=row, column=11, value=org['code'])
                    sheet.cell(row=row, column=12, value=org['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=13, value=org['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=14, value=org['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=15, value=org['utilization_percentage'] / 100).number_format = '0.0%'
                    row += 1
                if len(period_data['organization_summaries']) > 1:
                    for col in range(1, 10):
                        sheet.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)
        workbook.save(response)
        return response

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data = context.get(self.context_object_name, [])
        if output_format == 'excel':
            response = self.generate_excel_report(report_data)
            if response:
                return response
            messages.error(self.request, _("خطا در تولید فایل اکسل."))
            return super().render_to_response(context, **response_kwargs)
        return super().render_to_response(context, **response_kwargs)

class APIOrganizationsForPeriodView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API: Getting organizations for period {period_pk}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            allocations = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values('organization__id', 'organization__name').annotate(
                total_allocated=Sum('allocated_amount')
            ).order_by('organization__name')
            organizations_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'total_allocated_formatted': f"{org['total_allocated'] or 0:,.0f}",
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                    kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']})
                } for org in allocations
            ]
            html_content = render_to_string(
                'reports/partials/_ajax_level_organizations.html',
                {'organizations': organizations_data}
            )
            logger.debug(f"Rendered HTML content length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)
        except Http404:
            logger.error(f"404 Error: BudgetPeriod ID={period_pk} not found")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("دوره بودجه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} for Period PK={period_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش سازمان‌ها.")}),
                 'status': 'error'},
                status=500
            )

class BudgetItemsForOrgPeriodAPIView(APIView):
    def get(self, request, period_pk, org_pk, *args, **kwargs):

        logger.info(f"API request: BudgetItemsForOrgPeriodAPIView | Period PK={period_pk}, Org PK={org_pk}, User={request.user}, Path={request.path}, Headers={dict(request.headers)}")
        try:
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.info(f"Found BudgetPeriod: {budget_period.name} (ID={budget_period.pk}), Organization: {organization.name} (ID={organization.pk})")
            allocations = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').annotate(
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='RETURN'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('budget_item__name')
            logger.debug(f"Found {allocations.count()} BudgetAllocations: {[alloc.pk for alloc in allocations]}")
            if not allocations.exists():
                logger.warning(f"No BudgetAllocations found for Period ID={period_pk}, Org ID={org_pk}")
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                               {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'}, status=200)
            response_data = []
            for alloc in allocations:
                net_consumed = alloc.consumed - alloc.returned
                remaining = alloc.allocated_amount - net_consumed
                alloc_name = alloc.budget_item.name or 'سرفصل نامشخص'
                logger.debug(f"Processing Allocation ID={alloc.pk}, Name={alloc_name}")
                tankhahs = Tankhah.objects.filter(
                    project_budget_allocation=alloc,
                    is_archived=False
                ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')
                logger.debug(f"Found {tankhahs.count()} Tankhahs for Allocation ID={alloc.pk}")
                tankhah_data = []
                for t in tankhahs:
                    factors = t.factors.all()
                    factor_data = [
                        {
                            'id': f.pk,
                            'number': f.number,
                            'amount_formatted': f"{f.amount or 0:,.0f}",
                            'status_display': f.get_status_display(),
                            'category_name': f.category.name if f.category else "-",
                            'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime(
                                '%Y/%m/%d') if f.date else "-",
                            'detail_url': reverse('factor_detail', kwargs={'pk': f.pk}),
                            'items_count': f.items.count(),
                            'items': [
                                {
                                    'description': item.description,
                                    'amount_formatted': f"{item.amount or 0:,.0f}",
                                    'quantity': item.quantity,
                                    'unit_price': f"{item.unit_price or 0:,.0f}",
                                    'status_display': item.get_status_display(),
                                } for item in f.items.all()
                            ]
                        } for f in factors
                    ]
                    tankhah_data.append({
                        'id': t.pk,
                        'number': t.number,
                        'amount_formatted': f"{t.amount or 0:,.0f}",
                        'status_display': t.get_status_display(),
                        'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime(
                            '%Y/%m/%d') if t.date else '-',
                        'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                        'factors': factor_data,
                        'factors_count': len(factor_data),
                        'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    })
                response_data.append({
                    'ba_pk': alloc.pk,
                    'budget_item_name': alloc_name,
                    'budget_item_code': alloc.budget_item.code or '-',
                    'ba_allocated_amount_formatted': f"{alloc.allocated_amount or 0:,.0f}",
                    'ba_consumed_amount_formatted': f"{net_consumed or 0:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining or 0:,.0f}",
                    'ba_utilization_percentage': (net_consumed / alloc.allocated_amount * 100) if alloc.allocated_amount else 0,
                    'tankhahs': tankhah_data,
                    'tankhahs_count': len(tankhah_data),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_allocation', kwargs={'alloc_pk': alloc.pk}),
                    'add_tankhah_url': '',  # موقتاً خالی
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': alloc.pk}),
                })
            logger.info(f"Prepared {len(response_data)} BudgetAllocation records")
            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {'budget_allocations_data': response_data, 'parent_period_pk': period_pk, 'parent_org_pk': org_pk}
            )
            logger.debug(f"Rendered HTML content, length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)
        except Http404 as e:
            logger.error(f"404 Error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': str(e)}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                 {'error_message': _("خطا در پردازش سرفصل‌ها.")}),
                 'status': 'error'},
                status=500
            )

class APIFactorsForTankhahView(View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API: Getting factors for tankhah {tankhah_pk}")
        try:
            tankhah = get_object_or_404(Tankhah, pk=tankhah_pk, is_archived=False)
            factors = Factor.objects.filter(tankhah=tankhah).order_by('-date')
            factors_data = [
                {
                    'id': f.pk,
                    'number': f.number,
                    'amount_formatted': f"{f.amount or 0:,.0f}",
                    'status_display': f.get_status_display(),
                    'category_name': f.category.name if f.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime('%Y/%m/%d') if f.date else "-",
                    'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                    'items_count': f.items.count(),
                    'items': [
                        {
                            'description': item.description,
                            'amount_formatted': f"{item.amount or 0:,.0f}",
                            'quantity': item.quantity,
                            'unit_price': f"{item.unit_price or 0:,.0f}",
                            'status_display': item.get_status_display(),
                        } for item in f.items.all()
                    ]
                } for f in factors
            ]
            html_content = render_to_string(
                'reports/partials/_ajax_level_factors.html',
                {'factors_list': factors_data}
            )
            logger.debug(f"Rendered HTML content length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)
        except Http404:
            logger.error(f"404 Error: Tankhah ID={tankhah_pk} not found")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تنخواه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} for Tankhah PK={tankhah_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش فاکتورها.")}),
                 'status': 'error'},
                status=500
            )

class APITankhahsForAllocationView(View):
    def get(self, request, alloc_pk, *args, **kwargs):
        try:
            logger.info(f"Fetching tankhahs for allocation {alloc_pk}")

            allocation = get_object_or_404(
                BudgetAllocation.objects.select_related('budget_item'),
                pk=alloc_pk,
                is_active=True
            )

            tankhahs = Tankhah.objects.filter(
                project_budget_allocation=allocation,  # یا budget_allocation بسته به مدل شما
                is_archived=False
            ).select_related('created_by').prefetch_related('factors')

            # ایجاد یک لیست برای ذخیره داده‌های تنخواه‌ها
            tankhahs_data = []

            for t in tankhahs:
                # پردازش هر تنخواه و اضافه کردن به لیست
                tankhah_data = {
                    'id': t.pk,
                    'number': t.number,
                    'amount': t.amount,
                    'amount_formatted': f"{t.amount:,.0f}" if t.amount else "0",
                    'status': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                    # سایر فیلدهای مورد نیاز
                }
                tankhahs_data.append(tankhah_data)

            context = {
                'tankhahs': tankhahs_data,
                'allocation': {
                    'id': allocation.pk,
                    'name': allocation.budget_item.name
                }
            }

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                context
            )

            return JsonResponse({
                'html_content': html_content,
                'status': 'success'
            })

        except Exception as e:
            logger.error(f"Error in tankhah API: {str(e)}", exc_info=True)
            return JsonResponse({
                'html_content': render_to_string(
                    'reports/partials/_error_ajax.html',
                    {'error_message': "خطا در پردازش داده‌های تنخواه"}
                ),
                'status': 'error'
            }, status=500)


class OrganizationAllocationsAPIView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API Request: Get organization allocations for Period PK={period_pk}, User={request.user}, Path={request.path}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization_summaries = []
            org_alloc_queryset = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True)
            ).order_by('organization__name')
            for org_summary in org_alloc_queryset:
                consumed = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed = consumed - returned
                organization_summaries.append({
                    'id': org_summary['organization__id'],
                    'name': org_summary['organization__name'],
                    'code': org_summary['organization__code'],
                    'total_allocated_formatted': f"{org_summary['total_allocated_to_org']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{org_summary['total_allocated_to_org'] - net_consumed:,.0f}",
                    'num_budget_items': org_summary['num_budget_items'],
                    'utilization_percentage': (net_consumed / org_summary['total_allocated_to_org'] * 100) if
                    org_summary['total_allocated_to_org'] > 0 else 0,
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                    kwargs={'period_pk': period.pk, 'org_pk': org_summary['organization__id']})
                })
            logger.info(f"Prepared {len(organization_summaries)} Organization summaries")
            return JsonResponse({'organizations': organization_summaries, 'status': 'success'}, status=200)
        except Http404:
            logger.error(f"404 Error: BudgetPeriod ID={period_pk} not found")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("دوره بودجه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} for Period PK={period_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش سازمان‌ها.")}),
                 'status': 'error'},
                status=500
            )

class AComprehensiveBudgetReportView(PermissionBaseView, ListView):
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Starting get_queryset...")
        try:
            total_allocated_subquery = BudgetAllocation.objects.filter(
                budget_period_id=OuterRef('pk'), is_active=True
            ).values('budget_period_id').annotate(total=Sum('allocated_amount')).values('total')

            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='CONSUMPTION'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            returned_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='RETURN'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            queryset = BudgetPeriod.objects.filter(
                is_active=True, is_completed=False
            ).select_related('organization').annotate(
                period_total_allocated=Coalesce(Subquery(total_allocated_subquery), Decimal('0'),
                                                output_field=DecimalField()),
                period_total_consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('-start_date', 'name')

            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"Applying search filter: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)

            count = queryset.count()
            logger.info(f"Found {count} BudgetPeriods.")
            if count == 0:
                logger.warning("No active BudgetPeriods found.")
            return queryset
        except Exception as e:
            logger.error(f"Error in get_queryset: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در بارگذاری دوره‌های بودجه رخ داد."))
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        logger.info(f"[{self.__class__.__name__}] - Starting get_context_data...")
        context = super().get_context_data(**kwargs)
        periods = context.get('object_list', [])

        processed_data = []
        for period in periods:
            net_consumed = period.period_total_consumed - period.period_total_returned
            remaining = period.total_amount - net_consumed
            utilization = (net_consumed / period.total_amount * 100) if period.total_amount else 0

            org_summaries = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                org_total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                org_consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                org_returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='RETURN'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('organization__name')

            org_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'code': org['organization__code'],
                    'total_allocated': org['org_total_allocated'] or 0,
                    'total_allocated_formatted': f"{org['org_total_allocated'] or 0:,.0f}",
                    'net_consumed': (org['org_consumed'] - org['org_returned']) or 0,
                    'net_consumed_formatted': f"{(org['org_consumed'] - org['org_returned']) or 0:,.0f}",
                    'remaining': (org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0,
                    'remaining_formatted': f"{(org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0:,.0f}",
                    'num_budget_items': org['num_budget_items'],
                    'utilization_percentage': (
                            (org['org_consumed'] - org['org_returned']) / org['org_total_allocated'] * 100
                    ) if org['org_total_allocated'] else 0,
                    'budget_items_ajax_url': reverse(
                        'api_budget_items_for_org_period',
                        kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']}
                    )
                } for org in org_summaries
            ]

            processed_data.append({
                'period': period,
                'summary': {
                    'total_budget': period.total_amount or 0,
                    'total_allocated': period.period_total_allocated or 0,
                    'net_consumed': net_consumed or 0,
                    'remaining': remaining or 0,
                    'utilization_percentage': utilization,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime(
                        '%Y/%m/%d') if period.start_date else '-',
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime(
                        '%Y/%m/%d') if period.end_date else '-',
                },
                'organization_summaries': org_data
            })

        context[self.context_object_name] = processed_data
        context['report_main_title'] = _("گزارش جامع بودجه")
        context['current_search_period'] = self.request.GET.get('search_period', '')

        if not processed_data:
            messages.info(self.request, _("هیچ دوره بودجه فعالی یافت نشد."))
        return context

    def generate_excel_report(self, report_data):
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl is not installed.")
            messages.error(self.request, _("لطفاً openpyxl را نصب کنید."))
            return None

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="budget_report.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True

        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل (ریال)"),
            _("تخصیص (ریال)"), _("مصرف (ریال)"), _("مانده (ریال)"), _("% مصرف"),
            _("سازمان تخصیص"), _("کد"), _("تخصیص سازمان (ریال)"), _("مصرف سازمان (ریال)"),
            _("مانده سازمان (ریال)"), _("% مصرف سازمان")
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col)].width = 20

        row = 2
        for period_data in report_data:
            period = period_data['period']
            summary = period_data['summary']
            start_row = row

            if not period_data['organization_summaries']:
                sheet.cell(row=row, column=1, value=period.name)
                sheet.cell(row=row, column=2, value=period.organization.name)
                sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                row += 1
            else:
                for org in period_data['organization_summaries']:
                    sheet.cell(row=row, column=1, value=period.name)
                    sheet.cell(row=row, column=2, value=period.organization.name)
                    sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                    sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                    sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                    sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                    sheet.cell(row=row, column=10, value=org['name'])
                    sheet.cell(row=row, column=11, value=org['code'])
                    sheet.cell(row=row, column=12, value=org['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=13, value=org['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=14, value=org['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=15, value=org['utilization_percentage'] / 100).number_format = '0.0%'
                    row += 1

                if len(period_data['organization_summaries']) > 1:
                    for col in range(1, 10):
                        sheet.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)

        workbook.save(response)
        return response

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data = context.get(self.context_object_name, [])

        if output_format == 'excel':
            response = self.generate_excel_report(report_data)
            if response:
                return response
            messages.error(self.request, _("خطا در تولید فایل اکسل."))
            return super().render_to_response(context, **response_kwargs)

        return super().render_to_response(context, **response_kwargs)

# ===================================================================
# 2. ویوی API برای سازمان‌ها
# ===================================================================

# ===================================================================
# 3. ویوی API برای سرفصل‌ها و پروژه‌ها (مهم‌ترین ویو)
# ===================================================================
class APIBudgetItemsForOrgPeriodView(PermissionBaseView, View):
    """وظیفه: نمایش سرفصل‌های اصلی و پروژه‌های زیرمجموعه آن‌ها."""

    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API: Getting budget items for period {period_pk}, org {org_pk}")

        main_allocations = BudgetAllocation.objects.filter(
            budget_period_id=period_pk, organization_id=org_pk, project__isnull=True, is_active=True
        ).select_related('budget_item').order_by('budget_item__name')

        main_alloc_ids = main_allocations.values_list('pk', flat=True)
        project_allocations = BudgetAllocation.objects.filter(
            budget_allocation_id__in=main_alloc_ids, is_active=True
        ).select_related('project', 'subproject', 'budget_allocation')

        projects_by_parent = {parent_id: [] for parent_id in main_alloc_ids}
        for pba in project_allocations:
            projects_by_parent[pba.budget_allocation_id].append(pba)

        response_data = []
        for main_alloc in main_allocations:
            child_projects = [
                {
                    'pba_pk': pba.pk,
                    'project_name': f"{pba.project.name}{f' / {pba.subproject.name}' if pba.subproject else ''}",
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_pba', kwargs={'pba_pk': pba.pk})
                } for pba in projects_by_parent.get(main_alloc.pk, [])
            ]

            response_data.append({
                'budget_item_name': main_alloc.budget_item.name,
                'ba_pk': main_alloc.pk,
                'projects': child_projects
            })

        html_content = render_to_string(
            'reports/partials/_ajax_level_budget_items.html',
            {'allocations_data': response_data}
        )
        return JsonResponse({'html_content': html_content})

# ===================================================================
# 4. ویوی API برای تنخواه‌ها (فقط برای تخصیص پروژه)
# ===================================================================
class APITankhahsForPBAView(View):
    def get(self, request, pba_pk, *args, **kwargs):
        logger.info(f"API Request: Get tankhahs for Allocation PK={pba_pk}, User={request.user}, Path={request.path}")
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=pba_pk, is_active=True)
            logger.debug(f"Found BudgetAllocation ID={pba_pk}")

            tankhahs = Tankhah.objects.filter(
                budget_allocation=allocation,
                is_archived=False
            ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')
            logger.debug(f"Found {tankhahs.count()} Tankhahs: {[t.pk for t in tankhahs]}")

            if not tankhahs.exists():
                logger.warning(f"No tankhahs found for BudgetAllocation ID={pba_pk}")
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                logger.debug(f"Processing Tankhah ID={t.pk}, Number={t.number}")
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            logger.debug(f"Rendered HTML content length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            logger.error(f"404 Error: BudgetAllocation ID={pba_pk} not found")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تخصیص یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} for Allocation PK={pba_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                 'status': 'error'},
                status=500
            )

# ===================================================================
# 5. ویوی API برای فاکتورها
# ===================================================================
class  AA__APIFactorsForTankhahView(PermissionBaseView, View):
    """وظیفه: نمایش فاکتورهای یک تنخواه."""

    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API: Getting factors for tankhah {tankhah_pk}")
        factors = Factor.objects.filter(tankhah_id=tankhah_pk).order_by('-date')

        factors_data = [
            {
                'number': f.number,
                'amount_formatted': f"{f.amount:,.0f}"
            } for f in factors
        ]

        html_content = render_to_string(
            'reports/partials/_ajax_level_factors.html',
            {'factors_list': factors_data}
        )
        return JsonResponse({'html_content': html_content})
#---------------------------------------------------------
class  AA__BudgetItemsForOrgPeriodAPIView(View):
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request: BudgetItemsForOrgPeriodAPIView | Period PK={period_pk}, Org PK={org_pk}, User={request.user}, Path={request.path}, Headers={dict(request.headers)}")
        try:
            # دریافت دوره بودجه و سازمان
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.info(f"Found BudgetPeriod: {budget_period.name} (ID={budget_period.pk}), Organization: {organization.name} (ID={organization.pk})")

            # دریافت تخصیص‌های بودجه
            allocations = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').annotate(
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='RETURN'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('budget_item__name')

            logger.debug(f"Found {allocations.count()} BudgetAllocations: {[alloc.pk for alloc in allocations]}")

            if not allocations.exists():
                logger.warning(f"No BudgetAllocations found for Period ID={period_pk}, Org ID={org_pk}")
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                               {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'}, status=200)

            response_data = []
            for alloc in allocations:
                net_consumed = alloc.consumed - alloc.returned
                remaining = alloc.allocated_amount - net_consumed
                alloc_name = alloc.budget_item.name or 'سرفصل نامشخص'
                logger.debug(f"Processing Allocation ID={alloc.pk}, Name={alloc_name}")

                # تنخواه‌های متصل
                tankhahs = Tankhah.objects.filter(
                    budget_allocation=alloc,
                    is_archived=False
                ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')
                logger.debug(f"Found {tankhahs.count()} Tankhahs for Allocation ID={alloc.pk}")

                tankhah_data = []
                for t in tankhahs:
                    factors = t.factors.all()
                    factor_data = [
                        {
                            'id': f.pk,
                            'number': f.number,
                            'amount_formatted': f"{f.amount or 0:,.0f}",
                            'status_display': f.get_status_display(),
                            'category_name': f.category.name if f.category else "-",
                            'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime(
                                '%Y/%m/%d') if f.date else "-",
                            'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                            'items_count': f.items.count(),
                            'items': [
                                {
                                    'description': item.description,
                                    'amount_formatted': f"{item.amount or 0:,.0f}",
                                    'quantity': item.quantity,
                                    'unit_price': f"{item.unit_price or 0:,.0f}",
                                    'status_display': item.get_status_display(),
                                } for item in f.items.all()
                            ]
                        } for f in factors
                    ]

                    tankhah_data.append({
                        'id': t.pk,
                        'number': t.number,
                        'amount_formatted': f"{t.amount or 0:,.0f}",
                        'status_display': t.get_status_display(),
                        'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime(
                            '%Y/%m/%d') if t.date else '-',
                        'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                        'factors': factor_data,
                        'factors_count': len(factor_data),
                        'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    })

                response_data.append({
                    'ba_pk': alloc.pk,
                    'budget_item_name': alloc_name,
                    'budget_item_code': alloc.budget_item.code or '-',
                    'ba_allocated_amount_formatted': f"{alloc.allocated_amount or 0:,.0f}",
                    'ba_consumed_amount_formatted': f"{net_consumed or 0:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining or 0:,.0f}",
                    'ba_utilization_percentage': (net_consumed / alloc.allocated_amount * 100) if alloc.allocated_amount else 0,
                    'tankhahs': tankhah_data,
                    'tankhahs_count': len(tankhah_data),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_allocation', kwargs={'alloc_pk': alloc.pk}),
                    'add_tankhah_url': '',  # موقتاً خالی، چون tankhah_create تعریف نشده
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': alloc.pk}),
                })

            logger.info(f"Prepared {len(response_data)} BudgetAllocation records")
            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {'budget_allocations_data': response_data, 'parent_period_pk': period_pk, 'parent_org_pk': org_pk}
            )
            logger.debug(f"Rendered HTML content, length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)

        except Http404 as e:
            logger.error(f"404 Error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': str(e)}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                 {'error_message': _("خطا در پردازش سرفصل‌ها.")}),
                 'status': 'error'},
                status=500
            )

class  AzA__OrganizationAllocationsAPIView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization_summaries = []
            org_alloc_queryset = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            for org_summary in org_alloc_queryset:
                # محاسبه مصرف (این بخش همچنان نیاز به بهینه‌سازی دارد اگر تعداد زیاد است)
                consumed = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed = consumed - returned

                organization_summaries.append({
                    'id': org_summary['organization__id'],
                    'name': org_summary['organization__name'],
                    'code': org_summary['organization__code'],
                    'total_allocated_formatted': f"{org_summary['total_allocated_to_org']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{org_summary['total_allocated_to_org'] - net_consumed:,.0f}",
                    'num_budget_items': org_summary['num_budget_items'],
                    'utilization_percentage': (net_consumed / org_summary['total_allocated_to_org'] * 100) if
                    org_summary['total_allocated_to_org'] > 0 else 0,
                    # URL برای بارگذاری سرفصل‌های این سازمان
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_summary['organization__id']})
                })
            return JsonResponse({'organizations': organization_summaries, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("دوره بودجه یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (OrgAllocations) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش."), 'status': 'error'}, status=500)
class AbPITankhahsForAllocationView(View):  # PermissionBaseView حذف شد
    def get(self, request, alloc_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت تنخواه‌ها برای تخصیص PK={alloc_pk}")
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=alloc_pk, is_active=True)
            tankhahs = Tankhah.objects.filter(
                budget_allocation=allocation,
                is_archived=False
            ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')

            if not tankhahs.exists():
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تخصیص یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {str(e)}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                 'status': 'error'},
                status=500
            )

class AA__APIOrganizationsForPeriodView(PermissionBaseView, View):
    """وظیفه: فقط لیست سازمان‌های یک دوره را برمی‌گرداند."""

    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API: Getting organizations for period {period_pk}")
        period = get_object_or_404(BudgetPeriod, pk=period_pk)

        allocations = BudgetAllocation.objects.filter(
            budget_period=period, is_active=True, project__isnull=True
        ).values('organization__id', 'organization__name').annotate(
            total_allocated=Sum('allocated_amount')
        ).order_by('organization__name')

        organizations_data = [
            {
                'id': org['organization__id'],
                'name': org['organization__name'],
                'total_allocated_formatted': f"{org['total_allocated'] or 0:,.0f}",
                'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                 kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']})
            } for org in allocations
        ]

        html_content = render_to_string(
            'reports/partials/_ajax_level_organizations.html',
            {'organizations': organizations_data}
        )
        return JsonResponse({'html_content': html_content})


def get_factors_ajax(request, tankhah_id):
    tankhah = get_object_or_404(Tankhah, id=tankhah_id)
    factors = tankhah.factors.all()

    context = {
        'factors': factors,
        'tankhah': tankhah
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        html = render_to_string('reports/partials/factors_list.html', context, request)
        return JsonResponse({
            'success': True,
            'html': html
        })

    return JsonResponse({'success': False, 'message': 'Invalid request'})
