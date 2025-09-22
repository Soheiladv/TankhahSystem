"""
View های خروجی برای گزارش دسترسی کاربر
"""

import csv
import json
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.contrib.auth.mixins import UserPassesTestMixin
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from core.models import Transition, Post, Organization
from accounts.models import CustomUser
from core.models import UserPost
import logging

logger = logging.getLogger(__name__)


class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def handle_no_permission(self):
        return JsonResponse({'status': 'error', 'message': 'دسترسی غیرمجاز'}, status=403)


class ExportUserPermissionsView(StaffRequiredMixin, View):
    """خروجی Excel/CSV برای دسترسی‌های کاربر"""
    
    def get(self, request, user_id, format_type='csv'):
        user = get_object_or_404(CustomUser, pk=user_id)
        
        # پست‌های فعال کاربر
        user_post_ids = UserPost.objects.filter(user=user, is_active=True, post__is_active=True)\
                                       .values_list('post_id', flat=True)
        
        # ترنزیشن‌ها
        transitions = Transition.objects.filter(is_active=True)\
                        .select_related('entity_type', 'from_status', 'to_status', 'action', 'organization')\
                        .prefetch_related('allowed_posts').distinct()
        
        # محاسبه دسترسی
        user_post_set = set(user_post_ids)
        data = []
        
        for t in transitions:
            allowed_post_ids = set(t.allowed_posts.values_list('id', flat=True))
            has_access = bool(user_post_set & allowed_post_ids)
            
            data.append({
                'transition_id': t.id,
                'transition_name': t.name,
                'entity_type': t.entity_type.name,
                'from_status': t.from_status.name,
                'action': t.action.name,
                'to_status': t.to_status.name,
                'organization': t.organization.name,
                'has_access': 'بله' if has_access else 'خیر',
                'access_status': 'در دسترس' if has_access else 'خارج از دسترس'
            })
        
        if format_type == 'csv':
            return self.export_csv(data, user)
        elif format_type == 'excel':
            return self.export_excel(data, user)
        else:
            return JsonResponse({'error': 'فرمت پشتیبانی نمی‌شود'}, status=400)
    
    def export_csv(self, data, user):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="user_permissions_{user.username}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # اضافه کردن BOM برای UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # هدر
        writer.writerow([
            'ردیف',
            'شناسه گذار',
            'نام گذار',
            'نوع موجودیت',
            'از وضعیت',
            'اقدام',
            'به وضعیت',
            'سازمان',
            'دسترسی',
            'وضعیت دسترسی'
        ])
        
        # داده‌ها
        for i, row in enumerate(data, 1):
            writer.writerow([
                i,
                row['transition_id'],
                row['transition_name'],
                row['entity_type'],
                row['from_status'],
                row['action'],
                row['to_status'],
                row['organization'],
                row['has_access'],
                row['access_status']
            ])
        
        return response
    
    def export_excel(self, data, user):
        # برای Excel نیاز به کتابخانه openpyxl است
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return JsonResponse({'error': 'کتابخانه openpyxl نصب نیست'}, status=500)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"دسترسی‌های {user.username}"
        
        # استایل‌ها
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # هدر
        headers = [
            'ردیف', 'شناسه گذار', 'نام گذار', 'نوع موجودیت',
            'از وضعیت', 'اقدام', 'به وضعیت', 'سازمان',
            'دسترسی', 'وضعیت دسترسی'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # داده‌ها
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=row_data['transition_id'])
            ws.cell(row=row_idx, column=3, value=row_data['transition_name'])
            ws.cell(row=row_idx, column=4, value=row_data['entity_type'])
            ws.cell(row=row_idx, column=5, value=row_data['from_status'])
            ws.cell(row=row_idx, column=6, value=row_data['action'])
            ws.cell(row=row_idx, column=7, value=row_data['to_status'])
            ws.cell(row=row_idx, column=8, value=row_data['organization'])
            ws.cell(row=row_idx, column=9, value=row_data['has_access'])
            ws.cell(row=row_idx, column=10, value=row_data['access_status'])
        
        # تنظیم عرض ستون‌ها
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="user_permissions_{user.username}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response


class GenerateRuleTextView(StaffRequiredMixin, View):
    """تولید متن قانون برای کپی"""
    
    def get(self, request, user_id):
        user = get_object_or_404(CustomUser, pk=user_id)
        
        # پست‌های فعال کاربر
        user_post_ids = UserPost.objects.filter(user=user, is_active=True, post__is_active=True)\
                                       .values_list('post_id', flat=True)
        
        # ترنزیشن‌ها
        transitions = Transition.objects.filter(is_active=True)\
                        .select_related('entity_type', 'from_status', 'to_status', 'action', 'organization')\
                        .prefetch_related('allowed_posts').distinct()
        
        # محاسبه دسترسی
        user_post_set = set(user_post_ids)
        accessible_transitions = []
        inaccessible_transitions = []
        
        for t in transitions:
            allowed_post_ids = set(t.allowed_posts.values_list('id', flat=True))
            has_access = bool(user_post_set & allowed_post_ids)
            
            transition_info = {
                'name': t.name,
                'entity_type': t.entity_type.name,
                'from_status': t.from_status.name,
                'action': t.action.name,
                'to_status': t.to_status.name,
                'organization': t.organization.name
            }
            
            if has_access:
                accessible_transitions.append(transition_info)
            else:
                inaccessible_transitions.append(transition_info)
        
        # تولید متن قانون
        rule_text = f"""قانون دسترسی برای کاربر: {user.get_full_name() or user.username}
نام کاربری: {user.username}
تاریخ تولید: {timezone.now().strftime('%Y/%m/%d %H:%M')}

دسترسی‌های کاربر:
"""
        
        for i, t in enumerate(accessible_transitions, 1):
            rule_text += f"{i}. {t['name']} ({t['entity_type']}): از {t['from_status']} به {t['to_status']} با اقدام {t['action']} در سازمان {t['organization']}\n"
        
        if inaccessible_transitions:
            rule_text += f"\nعدم دسترسی‌های کاربر:\n"
            for i, t in enumerate(inaccessible_transitions, 1):
                rule_text += f"{i}. {t['name']} ({t['entity_type']}): از {t['from_status']} به {t['to_status']} با اقدام {t['action']} در سازمان {t['organization']}\n"
        
        return JsonResponse({
            'rule_text': rule_text,
            'accessible_count': len(accessible_transitions),
            'inaccessible_count': len(inaccessible_transitions),
            'total_count': len(transitions)
        })


class CopyUserPermissionsView(StaffRequiredMixin, View):
    """کپی دسترسی‌های یک کاربر به کاربر دیگر"""
    
    def post(self, request):
        try:
            source_user_id = request.POST.get('source_user_id')
            target_user_id = request.POST.get('target_user_id')
            copy_mode = request.POST.get('copy_mode', 'all')  # all, accessible_only, inaccessible_only
            
            if not source_user_id or not target_user_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'کاربر مبدأ و مقصد باید انتخاب شوند'
                }, status=400)
            
            if source_user_id == target_user_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'کاربر مبدأ و مقصد نمی‌توانند یکسان باشند'
                }, status=400)
            
            source_user = get_object_or_404(CustomUser, pk=source_user_id)
            target_user = get_object_or_404(CustomUser, pk=target_user_id)
            
            # دریافت پست‌های فعال کاربر مبدأ
            source_user_post_ids = UserPost.objects.filter(
                user=source_user, is_active=True, post__is_active=True
            ).values_list('post_id', flat=True)
            
            # دریافت پست‌های فعال کاربر مقصد
            target_user_post_ids = UserPost.objects.filter(
                user=target_user, is_active=True, post__is_active=True
            ).values_list('post_id', flat=True)
            
            # دریافت تمام ترنزیشن‌های فعال
            transitions = Transition.objects.filter(is_active=True)\
                            .prefetch_related('allowed_posts').distinct()
            
            # محاسبه دسترسی‌های کاربر مبدأ
            source_user_post_set = set(source_user_post_ids)
            target_user_post_set = set(target_user_post_ids)
            
            copied_count = 0
            skipped_count = 0
            
            for transition in transitions:
                allowed_post_ids = set(transition.allowed_posts.values_list('id', flat=True))
                source_has_access = bool(source_user_post_set & allowed_post_ids)
                
                # تصمیم‌گیری بر اساس حالت کپی
                should_copy = False
                if copy_mode == 'all':
                    should_copy = True
                elif copy_mode == 'accessible_only' and source_has_access:
                    should_copy = True
                elif copy_mode == 'inaccessible_only' and not source_has_access:
                    should_copy = True
                
                if should_copy:
                    if source_has_access:
                        # اضافه کردن دسترسی به کاربر مقصد
                        transition.allowed_posts.add(*target_user_post_ids)
                        copied_count += 1
                    else:
                        # حذف دسترسی از کاربر مقصد
                        transition.allowed_posts.remove(*target_user_post_ids)
                        copied_count += 1
                else:
                    skipped_count += 1
            
            # پاک کردن کش
            cache.delete(f"user_transitions_{target_user_id}")
            
            return JsonResponse({
                'status': 'success',
                'message': f'دسترسی‌های {source_user.get_full_name() or source_user.username} با موفقیت به {target_user.get_full_name() or target_user.username} کپی شد',
                'copied_count': copied_count,
                'skipped_count': skipped_count,
                'copy_mode': copy_mode
            })
            
        except Exception as e:
            logger.error(f"Error copying user permissions: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'خطا در کپی دسترسی‌ها: {str(e)}'
            }, status=500)


class GetUserPermissionsSummaryView(StaffRequiredMixin, View):
    """دریافت خلاصه دسترسی‌های کاربر"""
    
    def get(self, request, user_id):
        user = get_object_or_404(CustomUser, pk=user_id)
        
        # پست‌های فعال کاربر
        user_post_ids = UserPost.objects.filter(user=user, is_active=True, post__is_active=True)\
                                       .values_list('post_id', flat=True)
        
        # ترنزیشن‌ها
        transitions = Transition.objects.filter(is_active=True)\
                        .select_related('entity_type', 'from_status', 'to_status', 'action', 'organization')\
                        .prefetch_related('allowed_posts').distinct()
        
        # محاسبه دسترسی
        user_post_set = set(user_post_ids)
        accessible_count = 0
        inaccessible_count = 0
        
        for t in transitions:
            allowed_post_ids = set(t.allowed_posts.values_list('id', flat=True))
            has_access = bool(user_post_set & allowed_post_ids)
            
            if has_access:
                accessible_count += 1
            else:
                inaccessible_count += 1
        
        return JsonResponse({
            'user_id': user.id,
            'username': user.username,
            'full_name': user.get_full_name() or user.username,
            'accessible_count': accessible_count,
            'inaccessible_count': inaccessible_count,
            'total_count': len(transitions)
        })
