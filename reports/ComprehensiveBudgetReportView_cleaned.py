# reports/ComprehensiveBudgetReportView_cleaned.py
"""
نسخه پاکسازی شده ComprehensiveBudgetReportView
فقط API های فعال و ضروری
"""
import logging
from decimal import Decimal
import jdatetime

from django.db.models import Sum, Q, Count, OuterRef, Subquery, DecimalField, Prefetch
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, View
from django.contrib import messages
from rest_framework.views import APIView

# مدل‌های برنامه
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from core.PermissionBase import PermissionBaseView
from core.models import Organization, Project
from tankhah.models import Tankhah, Factor

# توابع محاسباتی
from budgets.budget_calculations import get_tankhah_used_budget

logger = logging.getLogger(__name__)

# ===================================================================
# 1. گزارش جامع بودجه - کلاس اصلی
# ===================================================================
class ComprehensiveBudgetReportView(PermissionBaseView, ListView):
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5
    permission_codes = ['BudgetPeriod.budgetperiod_reports']

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Starting get_queryset...")
        try:
            total_allocated_subquery = BudgetAllocation.objects.filter(
                budget_period_id=OuterRef('pk'), is_active=True
            ).values('budget_period_id').annotate(total=Sum('allocated_amount')).values('total')

            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='CONSUMPTION'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            returned_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='RETURN'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            queryset = BudgetPeriod.objects.filter(
                is_active=True, is_completed=False
            ).select_related('organization').annotate(
                period_total_allocated=Coalesce(Subquery(total_allocated_subquery), Decimal('0'),
                                                output_field=DecimalField()),
                period_total_consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('-start_date', 'name')

            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"Applying search filter: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)

            count = queryset.count()
            logger.info(f"Found {count} BudgetPeriods.")
            if count == 0:
                logger.warning("No active BudgetPeriods found.")
            return queryset
        except Exception as e:
            logger.error(f"Error in get_queryset: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در بارگذاری دوره‌های بودجه رخ داد."))
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        """آماده‌سازی داده‌های context برای نمایش"""
        logger.info(f"[{self.__class__.__name__}] - Starting get_context_data...")
        
        # دریافت queryset
        periods = self.get_queryset()
        
        # تنظیم object_list برای ListView
        self.object_list = periods
        
        # ایجاد context دستی
        context = {
            'object_list': periods,
            'is_paginated': False,
            'paginator': None,
            'page_obj': None,
        }
        
        processed_data = []
        
        for period in periods:
            try:
                net_consumed = period.period_total_consumed - period.period_total_returned
                remaining = period.total_amount - net_consumed
                utilization = (net_consumed / period.total_amount * 100) if period.total_amount else 0

                org_summaries = BudgetAllocation.objects.filter(
                    budget_period=period, is_active=True
                ).values(
                    'organization__id', 'organization__name', 'organization__code'
                ).annotate(
                    org_total_allocated=Sum('allocated_amount'),
                    num_budget_items=Count('budget_item', distinct=True),
                    org_consumed=Coalesce(
                        Subquery(
                            BudgetTransaction.objects.filter(
                                allocation__budget_period=period,
                                allocation__organization_id=OuterRef('organization__id'),
                                transaction_type='CONSUMPTION'
                            ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                        ), Decimal('0'), output_field=DecimalField()
                    ),
                    org_returned=Coalesce(
                        Subquery(
                            BudgetTransaction.objects.filter(
                                allocation__budget_period=period,
                                allocation__organization_id=OuterRef('organization__id'),
                                transaction_type='RETURN'
                            ).values('allocation__organization__id').annotate(total=Sum('amount')).values('total')
                        ), Decimal('0'), output_field=DecimalField()
                    )
                ).order_by('organization__name')

                org_data = [
                    {
                        'id': org['organization__id'],
                        'name': org['organization__name'],
                        'code': org['organization__code'],
                        'total_allocated': org['org_total_allocated'] or 0,
                        'total_allocated_formatted': f"{org['org_total_allocated'] or 0:,.0f}",
                        'net_consumed': (org['org_consumed'] - org['org_returned']) or 0,
                        'net_consumed_formatted': f"{(org['org_consumed'] - org['org_returned']) or 0:,.0f}",
                        'remaining': (org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0,
                        'remaining_formatted': f"{(org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0:,.0f}",
                        'num_budget_items': org['num_budget_items'],
                        'utilization_percentage': (
                            (org['org_consumed'] - org['org_returned']) / org['org_total_allocated'] * 100
                        ) if org['org_total_allocated'] else 0,
                        'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                         kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']})
                    } for org in org_summaries
                ]
                logger.info(f'org_summaries IS : {org_summaries}')
                processed_data.append({
                    'period': period,
                    'summary': {
                        'total_budget': period.total_amount or 0,
                        'total_allocated': period.period_total_allocated or 0,
                        'net_consumed': net_consumed or 0,
                        'remaining': remaining or 0,
                        'utilization_percentage': utilization,
                        'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime(
                            '%Y/%m/%d') if period.start_date else '-',
                        'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime(
                            '%Y/%m/%d') if period.end_date else '-',
                    },
                    'organization_summaries': org_data
                })
            except Exception as e:
                logger.error(f"Error processing period {period.pk}: {e}", exc_info=True)
                messages.error(self.request, _(f"خطا در پردازش دوره بودجه {period.name}."))

        context[self.context_object_name] = processed_data
        context['report_main_title'] = _("گزارش جامع بودجه")
        context['current_search_period'] = self.request.GET.get('search_period', '')

        if not processed_data:
            messages.info(self.request, _("هیچ دوره بودجه فعالی یافت نشد."))
        return context

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data = context.get(self.context_object_name, [])

        if output_format == 'excel':
            response = self.generate_excel_report(report_data)
            if response:
                return response
            messages.error(self.request, _("خطا در تولید فایل اکسل."))
            return super().render_to_response(context, **response_kwargs)

        return super().render_to_response(context, **response_kwargs)

    def generate_excel_report(self, report_data):
        """تولید گزارش Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش جامع بودجه"
            
            # هدرها
            headers = ['دوره بودجه', 'سازمان', 'کل بودجه', 'تخصیص یافته', 'مصرف شده', 'مانده', 'درصد مصرف']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # داده‌ها
            row = 2
            for period_data in report_data:
                period = period_data['period']
                for org_data in period_data['organization_summaries']:
                    ws.cell(row=row, column=1, value=period.name)
                    ws.cell(row=row, column=2, value=org_data['name'])
                    ws.cell(row=row, column=3, value=period_data['summary']['total_budget'])
                    ws.cell(row=row, column=4, value=org_data['total_allocated'])
                    ws.cell(row=row, column=5, value=org_data['net_consumed'])
                    ws.cell(row=row, column=6, value=org_data['remaining'])
                    ws.cell(row=row, column=7, value=org_data['utilization_percentage'])
                    row += 1
            
            # تنظیم عرض ستون‌ها
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="comprehensive_budget_report_{jdatetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return None

# ===================================================================
# 2. API سازمان‌ها برای یک دوره
# ===================================================================
class APIOrganizationsForPeriodView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API: Getting organizations for period {period_pk}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            
            organizations = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated=Sum('allocated_amount'),
                num_allocations=Count('id')
            ).order_by('organization__name')
            
            html_content = render_to_string(
                'reports/partials/_organizations_list.html',
                {'organizations': organizations, 'period': period}
            )
            
            return JsonResponse({'html_content': html_content})
            
        except Exception as e:
            logger.error(f"Error in APIOrganizationsForPeriodView: {e}")
            return JsonResponse({'error': str(e)}, status=500)

# ===================================================================
# 3. API سرفصل‌ها برای سازمان و دوره
# ===================================================================
class BudgetItemsForOrgPeriodAPIView(APIView):
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request: BudgetItemsForOrgPeriodAPIView | Period PK={period_pk}, Org PK={org_pk}")
        
        try:
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            
            logger.info(f"Found BudgetPeriod: {budget_period.name}, Organization: {organization.name}")

            # دریافت تخصیص‌های بودجه
            allocations = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').annotate(
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation_id').annotate(
                            total=Sum('amount')
                        ).values('total')
                    ), 
                    Decimal('0'), 
                    output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='RETURN'
                        ).values('allocation_id').annotate(
                            total=Sum('amount')
                        ).values('total')
                    ), 
                    Decimal('0'), 
                    output_field=DecimalField()
                )
            ).order_by('budget_item__name')

            logger.debug(f"Found {allocations.count()} BudgetAllocations")

            if not allocations.exists():
                logger.warning(f"No BudgetAllocations found for Period ID={period_pk}, Org ID={org_pk}")
                html_content = render_to_string(
                    'reports/partials/_no_budget_items_found.html',
                    {'organization': organization, 'period': budget_period}
                )
                return JsonResponse({'html_content': html_content, 'status': 'empty'}, status=200)

            response_data = []
            for alloc in allocations:
                net_consumed = alloc.consumed - alloc.returned
                remaining = alloc.allocated_amount - net_consumed
                alloc_name = alloc.budget_item.name or 'سرفصل نامشخص'

                # دریافت تنخواه‌های مرتبط
                tankhahs = Tankhah.objects.filter(
                    project_budget_allocation=alloc,
                    is_archived=False
                ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')

                tankhah_data = []
                for t in tankhahs:
                    factors = t.factors.all()
                    factor_data = [
                        {
                            'id': f.pk,
                            'number': f.number,
                            'amount_formatted': f"{f.amount or 0:,.0f}",
                            'status_display': f.status.name if f.status else 'نامشخص',
                            'category_name': f.category.name if f.category else "-",
                            'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime('%Y/%m/%d') if f.date else "-",
                            'detail_url': reverse('advance_factor_detail', kwargs={'factor_pk': f.pk}),
                            'items_count': f.items.count(),
                        } for f in factors
                    ]
                    tankhah_data.append({
                        'id': t.pk,
                        'number': t.number,
                        'amount_formatted': f"{t.amount or 0:,.0f}",
                        'status_display': t.status.name if t.status else 'نامشخص',
                        'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else '-',
                        'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                        'factors': factor_data,
                        'factors_count': len(factor_data),
                        'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    })

                response_data.append({
                    'ba_pk': alloc.pk,
                    'budget_item_name': alloc_name,
                    'budget_item_code': alloc.budget_item.code if alloc.budget_item else '-',
                    'ba_allocated_amount': alloc.allocated_amount,
                    'ba_allocated_amount_formatted': f"{alloc.allocated_amount:,.0f}",
                    'ba_consumed_amount': net_consumed,
                    'ba_consumed_amount_formatted': f"{net_consumed:,.0f}",
                    'ba_remaining_amount': remaining,
                    'ba_remaining_amount_formatted': f"{remaining:,.0f}",
                    'ba_utilization_percentage': (net_consumed / alloc.allocated_amount * 100) if alloc.allocated_amount else 0,
                    'tankhahs_count': tankhahs.count(),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_allocation', kwargs={'alloc_pk': alloc.pk}),
                    'tankhahs_data': tankhah_data,
                    'ba_report_url': '#'
                })

            logger.info(f"Prepared {len(response_data)} BudgetAllocation records")
            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {
                    'budget_allocations_data': response_data,
                    'organization': organization,
                    'period': budget_period
                }
            )

            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)

        except Http404 as e:
            logger.error(f"404 Error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}")
            return JsonResponse(
                {
                    'html_content': render_to_string(
                        'reports/partials/_no_data_found_ajax.html', 
                        {'message': str(e)}
                    ),
                    'status': 'notfound'
                },
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} | Period PK={period_pk}, Org PK={org_pk}", exc_info=True)
            return JsonResponse(
                {
                    'html_content': render_to_string(
                        'reports/partials/_error_ajax.html',
                        {'error_message': _("خطا در پردازش سرفصل‌ها.")}
                    ),
                    'status': 'error'
                },
                status=500
            )

# ===================================================================
# 4. API فاکتورها برای تنخواه
# ===================================================================
class APIFactorsForTankhahView(View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API: Getting factors for tankhah {tankhah_pk}")
        try:
            tankhah = get_object_or_404(Tankhah, pk=tankhah_pk, is_archived=False)
            
            factors = Factor.objects.filter(tankhah=tankhah).select_related('category', 'status').order_by('-date')
            
            factors_data = []
            for f in factors:
                factors_data.append({
                    'id': f.pk,
                    'number': f.number,
                    'amount_formatted': f"{f.amount or 0:,.0f}",
                    'status_display': f.status.name if f.status else 'نامشخص',
                    'category_name': f.category.name if f.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime('%Y/%m/%d') if f.date else "-",
                    'detail_url': reverse('advance_factor_detail', kwargs={'factor_pk': f.pk}),
                    'items_count': f.items.count(),
                })
            
            html_content = render_to_string(
                'reports/partials/_ajax_level_factors.html',
                {'factors_list': factors_data}
            )
            
            return JsonResponse({'html_content': html_content})
            
        except Exception as e:
            logger.error(f"Error in APIFactorsForTankhahView: {e}")
            return JsonResponse({'error': str(e)}, status=500)

# ===================================================================
# 5. API تنخواه‌ها برای تخصیص
# ===================================================================
class APITankhahsForAllocationView(View):
    def get(self, request, alloc_pk, *args, **kwargs):
        logger.info(f"API Request: Get tankhahs for Allocation PK={alloc_pk}")
        
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=alloc_pk, is_active=True)
            logger.debug(f"Found BudgetAllocation ID={alloc_pk}")

            tankhahs = Tankhah.objects.filter(
                project_budget_allocation=allocation,
                is_archived=False
            ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')

            logger.debug(f"Found {tankhahs.count()} Tankhahs")

            if not tankhahs.exists():
                logger.warning(f"No tankhahs found for BudgetAllocation ID={alloc_pk}")
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                logger.debug(f"Processing Tankhah ID={t.pk}, Number={t.number}")
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.status.name if t.status else 'نامشخص',
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'}, status=200)

        except Http404 as e:
            logger.error(f"404 Error: {str(e)} | Allocation PK={alloc_pk}")
            return JsonResponse(
                {
                    'html_content': render_to_string(
                        'reports/partials/_no_data_found_ajax.html', 
                        {'message': _("تخصیص یافت نشد.")}
                    ),
                    'status': 'notfound'
                },
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} | Allocation PK={alloc_pk}", exc_info=True)
            return JsonResponse(
                {
                    'html_content': render_to_string(
                        'reports/partials/_error_ajax.html', 
                        {'error_message': _("خطا در پردازش تنخواه‌ها.")}
                    ),
                    'status': 'error'
                },
                status=500
            )

# ===================================================================
# 6. API تخصیص‌های سازمان برای دوره
# ===================================================================
class OrganizationAllocationsAPIView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API Request: Get organization allocations for Period PK={period_pk}, User={request.user}, Path={request.path}")
        
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            
            organizations = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated=Sum('allocated_amount'),
                num_allocations=Count('id')
            ).order_by('organization__name')
            
            html_content = render_to_string(
                'reports/partials/_organizations_list.html',
                {'organizations': organizations, 'period': period}
            )
            
            return JsonResponse({'html_content': html_content})
            
        except Exception as e:
            logger.error(f"Error in OrganizationAllocationsAPIView: {e}")
            return JsonResponse({'error': str(e)}, status=500)
