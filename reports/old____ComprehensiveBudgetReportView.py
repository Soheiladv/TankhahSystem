# reports/views.py
import logging
from decimal import Decimal
import jdatetime
import logging
from decimal import Decimal
from django.urls import reverse_lazy  # برای لینک‌ها
from django.views.generic import ListView, View
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Prefetch, Q, Count, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce

# مدل‌های مورد نیاز
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetItem, ProjectBudgetAllocation, BudgetTransaction
from django.contrib.auth.mixins import LoginRequiredMixin  # جایگزین ساده

# reports/views.py
import logging
from decimal import Decimal
from django.urls import reverse, reverse_lazy  # reverse برای لینک‌های داخلی در context
from django.views.generic import ListView
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Prefetch, Q, Count, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, Http404
from django.template.loader import render_to_string
import jdatetime  # برای نمایش تاریخ شمسی

# مدل‌های مورد نیاز
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetItem, ProjectBudgetAllocation, BudgetTransaction
from core.models import Organization, Project, SubProject  # و سایر مدل‌های core که لازم است
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages  # برای نمایش پیام به کاربر

# توابع محاسباتی (فرض می‌کنیم اینها در دسترس و بهینه هستند)
import logging
from decimal import Decimal
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Prefetch, Q, Count, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.template.loader import render_to_string
import jdatetime

logger = logging.getLogger('comprehensive_report_v2_full')  # نام لاگر جدید

from django.shortcuts import get_object_or_404

# مدل‌های مورد نیاز
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetItem, ProjectBudgetAllocation, BudgetTransaction
from core.models import Organization, Project, SubProject
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages


# کتابخانه‌های PDF و Excel (مطمئن شوید نصب شده‌اند)
try:
    from weasyprint import HTML, CSS

    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger = logging.getLogger('comprehensive_report')
    logger.warning("WeasyPrint is not installed. PDF export will not be available.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger = logging.getLogger('comprehensive_report')
    logger.warning("openpyxl is not installed. Excel export will not be available.")


class old_ComprehensiveBudgetReportView(LoginRequiredMixin, ListView):
    model = BudgetPeriod
    template_name = 'reports/comprehensive_budget_report.html'
    context_object_name = 'budget_periods_report_data'  # نام context برای لیست دوره‌ها

    # paginate_by = 5 # اگر تعداد دوره‌های بودجه زیاد است، فعال کنید

    # permission_codenames = ['reports.view_comprehensive_budget_report'] # اگر از PermissionBaseView استفاده می‌کنید

    def get_queryset(self):
        # انتخاب دوره‌های بودجه فعال و تکمیل نشده (یا بر اساس فیلترهای دیگر)
        queryset = BudgetPeriod.objects.filter(
            is_active=True,
            is_completed=False
        ).select_related(
            'organization',  # سازمان اصلی که این دوره بودجه برای آن تعریف شده
            'created_by'
        ).annotate(
            # می‌توانید تعداد کل BudgetAllocation های مرتبط با هر دوره را هم اینجا annotate کنید
            num_allocations=Count('allocations', filter=Q(allocations__is_active=True))
            # related_name از BudgetPeriod به BudgetAllocation
        ).order_by('-start_date', 'name')  # جدیدترین‌ها و سپس بر اساس نام

        logger.info(f"[{self.__class__.__name__}] Fetching BudgetPeriods. Initial count: {queryset.count()}")
        # اینجا می‌توانید فیلترهای GET را اعمال کنید (مثلا برای جستجوی نام دوره یا سازمان)
        # search_query = self.request.GET.get('q', '')
        # if search_query:
        #     queryset = queryset.filter(Q(name__icontains=search_query) | Q(organization__name__icontains=search_query))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # object_list در اینجا شامل نمونه‌های BudgetPeriod از get_queryset است.

        report_data_for_template = []  # لیستی که به تمپلیت پاس داده می‌شود

        for period_instance in context['object_list']:
            logger.debug(f"Processing data for BudgetPeriod: {period_instance.name} (PK: {period_instance.pk})")

            # الف) خلاصه کلی خود دوره بودجه
            # کل تخصیص یافته مستقیم از این دوره به BudgetAllocation ها
            total_allocated_from_this_period = BudgetAllocation.objects.filter(
                budget_period=period_instance, is_active=True
            ).aggregate(
                total=Coalesce(Sum('allocated_amount'), Decimal('0'), output_field=DecimalField())
            )['total']

            # کل مصرف شده از تمام BudgetAllocation های این دوره
            # (این کوئری می‌تواند سنگین باشد، برای گزارش‌های پیچیده باید بهینه‌تر شود یا با سیگنال آپدیت شود)
            all_transactions_in_period = BudgetTransaction.objects.filter(allocation__budget_period=period_instance)
            consumed_in_period = all_transactions_in_period.filter(transaction_type='CONSUMPTION').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            returned_in_period = all_transactions_in_period.filter(transaction_type='RETURN').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            net_consumed_from_this_period = consumed_in_period - returned_in_period

            remaining_of_period_vs_total_amount = period_instance.total_amount - net_consumed_from_this_period  # مانده دوره نسبت به مصرف
            utilization_of_period_vs_total_amount = (net_consumed_from_this_period / period_instance.total_amount * 100) \
                if period_instance.total_amount > 0 else Decimal('0')

            # ب) اطلاعات تخصیص‌ها به سازمان‌ها از این دوره بودجه
            organization_summaries = []
            # دریافت BudgetAllocation های فعال برای این دوره، گروه‌بندی شده بر اساس سازمان
            # و محاسبه مجموع تخصیص و مصرف برای هر سازمان از این دوره
            # این کوئری باید بهینه باشد.
            org_alloc_queryset = BudgetAllocation.objects.filter(
                budget_period=period_instance, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'  # اطلاعات سازمان
            ).annotate(
                # مجموع تخصیص یافته به این سازمان از این دوره (از تمام سرفصل‌ها)
                total_allocated_to_org_from_this_period=Sum('allocated_amount'),
                # تعداد سرفصل‌های تخصیص یافته به این سازمان از این دوره
                num_budget_items_for_org=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            logger.debug(
                f"Found {org_alloc_queryset.count()} distinct organizations allocated from period {period_instance.name}")

            for org_summary in org_alloc_queryset:
                # برای هر سازمان، مجموع مصرف آن از این دوره بودجه را محاسبه می‌کنیم
                # (این کوئری در حلقه است، برای تعداد زیاد سازمان باید بهینه شود، مثلا با Subquery در annotate بالا)
                consumed_by_this_org_from_this_period = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

                returned_to_this_org_from_this_period = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

                net_consumed_by_this_org = consumed_by_this_org_from_this_period - returned_to_this_org_from_this_period

                remaining_for_this_org = org_summary[
                                             'total_allocated_to_org_from_this_period'] - net_consumed_by_this_org

                # یافتن اولین BudgetAllocation برای این سازمان از این دوره (برای لینک "نمونه گزارش سرفصل")
                first_ba_for_org_instance = BudgetAllocation.objects.filter(
                    budget_period=period_instance,
                    organization_id=org_summary['organization__id'],
                    is_active=True
                ).order_by('pk').first()  # یا هر ترتیب دیگری

                organization_summaries.append({
                    'id': org_summary['organization__id'],
                    'name': org_summary['organization__name'],
                    'code': org_summary['organization__code'],
                    'total_allocated': org_summary['total_allocated_to_org_from_this_period'],
                    'net_consumed': net_consumed_by_this_org,
                    'remaining': remaining_for_this_org,
                    'num_budget_items': org_summary['num_budget_items_for_org'],
                    'utilization_percentage': (
                                net_consumed_by_this_org / org_summary['total_allocated_to_org_from_this_period'] * 100) \
                        if org_summary['total_allocated_to_org_from_this_period'] > 0 else Decimal('0'),
                    # URL برای مشاهده تمام BudgetAllocation های این سازمان از این دوره
                    'allocations_list_url': reverse_lazy('budget_allocation_list_by_org_period',
                                                         # یک URL فرضی جدید
                                                         kwargs={'period_pk': period_instance.pk,
                                                                 'org_pk': org_summary['organization__id']}),
                    'first_ba_instance_pk_for_report_link': first_ba_for_org_instance.pk if first_ba_for_org_instance else None,
                })

            report_data_for_template.append({
                'period': period_instance,  # خود آبجکت دوره بودجه
                'summary': {
                    'total_budget': period_instance.total_amount,
                    'total_allocated': total_allocated_from_this_period,
                    'net_consumed': net_consumed_from_this_period,
                    'remaining_vs_consumed': remaining_of_period_vs_total_amount,
                    'utilization_percentage': utilization_of_period_vs_total_amount,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period_instance.start_date).strftime('%Y/%m/%d') if period_instance.start_date else "-",
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period_instance.end_date).strftime( '%Y/%m/%d') if period_instance.end_date else "-",
                },
                'organization_allocations': organization_summaries,  # لیست خلاصه تخصیص‌ها به سازمان‌ها
            })

        context[self.context_object_name] = report_data_for_template
        context['report_main_title'] = _("گزارش جامع بودجه و تخصیص‌ها")
        logger.info(
            f"[{self.__class__.__name__}] Context prepared. Processed {len(report_data_for_template)} budget periods.")
        return context


    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format')
        if output_format == 'pdf':
            # منطق تولید PDF با استفاده از داده‌های context
            # return generate_pdf_response(context) # یک تابع فرضی
            pass
        elif output_format == 'excel':
            # منطق تولید فایل Excel
            # return generate_excel_response(context) # یک تابع فرضی
            pass
        return super().render_to_response(context, **response_kwargs)


class old_2_ComprehensiveBudgetReportView(LoginRequiredMixin, ListView):
    model = BudgetPeriod
    template_name = 'reports/comprehensive_budget_report.html'
    context_object_name = 'budget_periods_report_data'

    # paginate_by = 5 # برای بارگذاری اولیه سبک، صفحه‌بندی در سطح دوره‌ها ممکن است لازم نباشد

    def get_queryset(self):
        queryset = BudgetPeriod.objects.filter(
            is_active=True,
            is_completed=False  # یا هر فیلتر اولیه دیگر
        ).select_related(
            'organization',  # سازمان اصلی که این دوره بودجه برای آن تعریف شده
        ).order_by('-start_date', 'name')

        # فیلترهای GET (مثال: جستجو بر اساس نام دوره)
        search_query = self.request.GET.get('search_period', '')
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)

        logger.info(f"[{self.__class__.__name__}] Fetching BudgetPeriods. Count: {queryset.count()}")
        return queryset

    def get_report_data_for_html(self):
        """
        داده‌ها را برای نمایش اولیه HTML آماده می‌کند (سطح اول و دوم).
        جزئیات بیشتر باید با AJAX بارگذاری شوند.
        """
        processed_periods = []
        # self.object_list توسط ListView از get_queryset پر می‌شود.
        # اگر get_queryset را override کرده‌اید و super() را صدا نمی‌زنید، باید خودتان self.object_list را مقداردهی کنید.
        # برای اطمینان، queryset را دوباره اینجا می‌گیریم.
        queryset = self.get_queryset()
        self.object_list = queryset  # برای سازگاری با ListView

        for period_instance in self.object_list:
            logger.debug(f"Processing HTML data for BudgetPeriod: {period_instance.name} (PK: {period_instance.pk})")

            # الف) خلاصه کلی خود دوره بودجه
            allocations_in_period = BudgetAllocation.objects.filter(budget_period=period_instance, is_active=True)
            total_allocated_from_this_period = allocations_in_period.aggregate(
                total=Coalesce(Sum('allocated_amount'), Decimal('0'), output_field=DecimalField())
            )['total']

            # مصرف واقعی از کل دوره (این کوئری می‌تواند سنگین باشد، به خصوص اگر تراکنش‌ها زیادند)
            all_transactions_in_period = BudgetTransaction.objects.filter(allocation__budget_period=period_instance)
            consumed_in_period = all_transactions_in_period.filter(transaction_type='CONSUMPTION').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            returned_in_period = all_transactions_in_period.filter(transaction_type='RETURN').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            net_consumed_from_this_period = consumed_in_period - returned_in_period

            remaining_of_period_vs_total_amount = period_instance.total_amount - net_consumed_from_this_period
            utilization_of_period_vs_total_amount = (net_consumed_from_this_period / period_instance.total_amount * 100) \
                if period_instance.total_amount > 0 else Decimal('0')

            # ب) خلاصه تخصیص‌ها به سازمان‌ها از این دوره بودجه
            organization_summaries = []
            # کوئری بهینه شده برای گرفتن مجموع تخصیص‌ها و تعداد سرفصل‌ها برای هر سازمان در این دوره
            org_alloc_queryset_summary = BudgetAllocation.objects.filter(
                budget_period=period_instance, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org_from_this_period=Sum('allocated_amount'),
                num_budget_items_for_org=Count('budget_item', distinct=True)
                # برای محاسبه مصرف هر سازمان به صورت بهینه، نیاز به Subquery یا OuterRef است
                # که در اینجا برای سادگی، مصرف را بعدا در حلقه (که N+1 ایجاد می‌کند) محاسبه می‌کنیم
                # یا در API جداگانه.
            ).order_by('organization__name')

            for org_summary_data in org_alloc_queryset_summary:
                # محاسبه مصرف برای این سازمان از این دوره (این بخش باید برای پروداکشن بهینه شود)
                consumed_by_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_summary_data['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned_to_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_summary_data['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed_by_this_org_val = consumed_by_this_org - returned_to_this_org

                remaining_for_this_org_val = org_summary_data[
                                                 'total_allocated_to_org_from_this_period'] - net_consumed_by_this_org_val

                # لینک برای بارگذاری جزئیات سرفصل‌ها و پروژه‌های این سازمان با AJAX
                ajax_detail_url = reverse('api_budget_items_for_org_period',  # نام URL جدید برای API
                                          kwargs={'period_pk': period_instance.pk,
                                                  'org_pk': org_summary_data['organization__id']})

                organization_summaries.append({
                    'id': org_summary_data['organization__id'],
                    'name': org_summary_data['organization__name'],
                    'code': org_summary_data['organization__code'],
                    'total_allocated': org_summary_data['total_allocated_to_org_from_this_period'],
                    'net_consumed': net_consumed_by_this_org_val,
                    'remaining': remaining_for_this_org_val,
                    'num_budget_items': org_summary_data['num_budget_items_for_org'],
                    'utilization_percentage': (net_consumed_by_this_org_val / org_summary_data[
                        'total_allocated_to_org_from_this_period'] * 100) \
                        if org_summary_data['total_allocated_to_org_from_this_period'] > 0 else Decimal('0'),
                    'ajax_detail_url': ajax_detail_url,  # URL برای AJAX
                    # لینک به گزارش BudgetAllocationReportView (جزئیات اولین/مهمترین سرفصل)
                    # این بخش هم می‌تواند با AJAX و پس از انتخاب سرفصل انجام شود
                    # 'first_ba_report_url': reverse_lazy('budget_allocation_report', kwargs={'pk': first_ba_pk_if_any})
                })

            processed_periods.append({
                'period': period_instance,
                'summary': {
                    'total_budget': period_instance.total_amount,
                    'total_allocated': total_allocated_from_this_period,
                    'net_consumed': net_consumed_from_this_period,
                    'remaining_vs_consumed': remaining_of_period_vs_total_amount,
                    'utilization_percentage': utilization_of_period_vs_total_amount,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period_instance.start_date).strftime(
                        '%Y/%m/%d') if period_instance.start_date else "-",
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period_instance.end_date).strftime(
                        '%Y/%m/%d') if period_instance.end_date else "-",
                },
                'organization_summaries': organization_summaries,
            })
        return processed_periods

    def get_context_data(self, **kwargs):
        # اطمینان از اینکه get_queryset قبل از get_report_data_for_html فراخوانی شده
        # ListView این کار را به صورت خودکار انجام می‌دهد و self.object_list را پر می‌کند.
        context = super().get_context_data(**kwargs)  # این self.object_list را به context اضافه می‌کند

        # حالا از self.object_list که توسط ListView مقداردهی شده، استفاده می‌کنیم
        # و آن را برای متد get_report_data_for_html که کمی تغییر دادیم، ارسال نمی‌کنیم.
        # یا اینکه get_report_data_for_html را طوری بنویسیم که object_list را از context بگیرد.
        # برای سادگی، get_report_data_for_html را طوری تغییر می‌دهیم که self.object_list را استفاده کند.

        context[self.context_object_name] = self.get_report_data_for_html()  # بازنویسی context_object_name
        context['report_main_title'] = _("گزارش جامع بودجه و تخصیص‌ها")
        logger.info(f"[{self.__class__.__name__}] HTML Context prepared.")
        return context

    def generate_excel_report(self, report_data):
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library is not installed. Cannot generate Excel report.")
            return None

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="comprehensive_budget_report.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True

        # استایل‌ها
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # آبی تیره
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_align_right = Alignment(horizontal='right', vertical='center')
        data_align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        currency_format = '#,##0'

        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل دوره"),
            _("کل تخصیص از دوره"), _("کل مصرف از دوره"), _("مانده دوره (به مصرف)"), _("% مصرف دوره"),
            _("سازمان تخصیص گیرنده"), _("کد سازمان"), _("تخصیص به سازمان"), _("مصرف سازمان از دوره"),
            _("مانده سازمان از دوره"), _("% مصرف سازمان")
        ]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            sheet.column_dimensions[get_column_letter(col_num)].width = 22

        row_num = 2
        for period_item_data in report_data:  # report_data همان budget_periods_report_data است
            period = period_item_data['period']
            p_summary = period_item_data['summary']

            start_row_for_period = row_num
            for org_sum_data in period_item_data['organization_summaries']:
                # اطلاعات دوره (برای هر ردیف سازمان تکرار می‌شود، مگر اینکه بعدا مرج کنیم)
                sheet.cell(row=row_num, column=1, value=period.name).border = thin_border
                sheet.cell(row=row_num, column=2, value=period.organization.name).border = thin_border
                sheet.cell(row=row_num, column=3, value=p_summary['start_date_jalali']).border = thin_border
                sheet.cell(row=row_num, column=4, value=p_summary['end_date_jalali']).border = thin_border
                cell = sheet.cell(row=row_num, column=5, value=p_summary['total_budget']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=6, value=p_summary['total_allocated']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=7, value=p_summary['net_consumed']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=8, value=p_summary['remaining_vs_consumed']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=9, value=p_summary['utilization_percentage'] / 100);
                cell.number_format = '0.0%';
                cell.border = thin_border

                # اطلاعات سازمان
                sheet.cell(row=row_num, column=10, value=org_sum_data['name']).border = thin_border
                sheet.cell(row=row_num, column=11, value=org_sum_data['code']).border = thin_border
                cell = sheet.cell(row=row_num, column=12, value=org_sum_data['total_allocated']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=13, value=org_sum_data['net_consumed']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=14, value=org_sum_data['remaining']);
                cell.number_format = currency_format;
                cell.border = thin_border
                cell = sheet.cell(row=row_num, column=15, value=org_sum_data['utilization_percentage'] / 100);
                cell.number_format = '0.0%';
                cell.border = thin_border
                row_num += 1

            # مرج کردن سلول‌های دوره بودجه اگر بیش از یک سازمان دارد
            if period_item_data['organization_summaries'] and len(period_item_data['organization_summaries']) > 1:
                for col_idx in range(1, 10):  # ستون‌های مربوط به دوره
                    sheet.merge_cells(start_row=start_row_for_period, start_column=col_idx, end_row=row_num - 1,
                                      end_column=col_idx)
                    sheet.cell(row=start_row_for_period, column=col_idx).alignment = Alignment(
                        vertical='top')  # تراز عمودی

            # اگر هیچ سازمانی برای دوره وجود نداشت، اطلاعات دوره را یکبار بنویس
            if not period_item_data['organization_summaries']:
                sheet.cell(row=row_num, column=1, value=period.name).border = thin_border
                sheet.cell(row=row_num, column=2, value=period.organization.name).border = thin_border
                # ... (بقیه فیلدهای دوره)
                row_num += 1

        try:
            workbook.save(response)
            return response
        except Exception as e:
            logger.error(f"Error saving Excel workbook: {e}", exc_info=True)
            return None  # یا یک HttpResponse با خطا

    def generate_pdf_report(self, report_data, report_title):
        if not WEASYPRINT_AVAILABLE:
            logger.error("WeasyPrint library is not installed. Cannot generate PDF report.")
            return None

        # استفاده از تمپلیت جداگانه برای PDF
        html_string = render_to_string(
            'reports/pdf/comprehensive_budget_report_pdf.html',  # این تمپلیت را باید بسازید
            {
                'budget_periods_report_data': report_data,
                'report_main_title': report_title
            }
        )
        # اضافه کردن استایل‌های CSS برای PDF
        # شما باید یک فایل CSS مناسب برای PDF ایجاد کنید یا استایل‌ها را به صورت رشته پاس دهید.
        # pdf_css = CSS(string="body { font-family: 'B Nazanin', 'Tahoma', sans-serif; } table { width: 100%; border-collapse: collapse;} th, td {border: 1px solid black; padding: 5px; text-align: right;}") # مثال خیلی ساده

        try:
            # برای فونت فارسی، باید مسیر فونت را به WeasyPrint بدهید یا از font-face در CSS استفاده کنید
            # font_config = FontConfiguration()
            # css_for_pdf = CSS(string='@font-face {font-family: Vazirmatn; src: url(/static/fonts/your-persian-font.ttf); } body { font-family: Vazirmatn; }', font_config=font_config)
            # html_doc = HTML(string=html_string, base_url=self.request.build_absolute_uri())
            # pdf_file = html_doc.write_pdf(stylesheets=[css_for_pdf]) # استفاده از استایل

            # راه ساده تر بدون تنظیمات فونت پیچیده (ممکن است فونت پیش‌فرض خوب نباشد)
            # یک تمپلیت HTML ساده شده برای PDF بسازید که فقط جداول و اطلاعات لازم را دارد.
            pdf_css_string_internal = """
                @page { size: A3 landscape; margin: 1cm; } /* A3 landscape برای جای بیشتر */
                body { font-family: "DejaVu Sans", "Arial", sans-serif; direction: rtl; font-size: 8pt; }
                table { width: 100%; border-collapse: collapse; margin-bottom: 12px; page-break-inside: auto; }
                th, td { border: 1px solid #777; padding: 3px; text-align: right; word-wrap: break-word; }
                th { background-color: #eaeaea; font-weight: bold; }
                h1, h2, h3 { text-align: center; margin-bottom: 8px; font-weight: bold; color: #333; }
                h1 { font-size: 14pt; } h2 { font-size: 11pt; } h3 { font-size: 9pt; }
                .period-summary-pdf { background-color: #f9f9f9; padding: 6px; margin-bottom:8px; border: 1px dashed #bbb; font-size: 7.5pt;}
                .org-summary-pdf { margin-right: 15px; margin-bottom: 6px; border: 1px solid #ddd; padding: 4px; font-size: 7.5pt;}
                .text-danger-pdf { color: #d00; } .text-success-pdf { color: #070; }
            """
            html_doc = HTML(string=html_string, base_url=self.request.build_absolute_uri())
            pdf_file = html_doc.write_pdf(stylesheets=[CSS(string=pdf_css_string_internal)])

            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="comprehensive_budget_report.pdf"'
            return response
        except Exception as e:
            logger.error(f"Error generating PDF with WeasyPrint: {e}", exc_info=True)
            return None

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data_for_export = context.get(self.context_object_name, [])
        report_title_for_export = context.get('report_main_title', _("گزارش بودجه"))

        if output_format == 'pdf':
            pdf_response = self.generate_pdf_report(report_data_for_export, report_title_for_export)
            if pdf_response:
                return pdf_response
            else:
                messages.error(self.request,
                               _("خطا در تولید خروجی PDF. لطفاً از نصب بودن WeasyPrint و وابستگی‌های آن اطمینان حاصل کنید."))

        elif output_format == 'excel':
            excel_response = self.generate_excel_report(report_data_for_export)
            if excel_response:
                return excel_response
            else:
                messages.error(self.request,
                               _("خطا در تولید خروجی اکسل. لطفاً از نصب بودن openpyxl اطمینان حاصل کنید."))

        # اگر فرمت HTML یا خطایی در خروجی دیگر رخ داده، تمپلیت اصلی را رندر کن
        # اطمینان از اینکه context['object_list'] برای ListView مقداردهی شده
        # اگر get_context_data را override کرده‌ایم، باید خودمان object_list را از queryset اصلی بگیریم
        # یا در get_context_data، context[self.context_object_name] را به object_list هم بدهیم
        # context['object_list'] = context.get(self.context_object_name) # این ممکن است برای paginate_by مشکل ایجاد کند
        # راه بهتر: ListView خودش object_list را از get_queryset می‌سازد و به context اضافه می‌کند
        # ما در get_context_data فقط context[self.context_object_name] را با داده پردازش شده خودمان جایگزین کردیم
        # پس باید super().get_context_data() را برای گرفتن context اولیه (شامل paginator و page_obj) فراخوانی کنیم.
        # اما چون paginate_by را کامنت کردیم، فعلا نیازی نیست.

        return super().render_to_response(context, **response_kwargs)
#
"""
شرح تک خطی: این ویو، نمای کلی گزارش جامع رو نمایش می‌ده و اولین لایه (دوره‌های بودجه) رو بارگذاری می‌کنه.

شرح تک خطی: این API، لیست سرفصل‌های بودجه و تخصیص‌های پروژه مرتبط (که از قبل واکشی شده‌اند) رو برای یک سازمان و دوره بودجه مشخص برمی‌گردونه.
تمپلیت مرتبط: reports/partials/_ajax_level_budget_allocations.html (این HTML رو در پاسخ JSON برمی‌گردونه).
نحوه فراخوانی API‌ها در این تمپلیت:
سطرهای سرفصل بودجه (ba): هر سطر <tr> برای سرفصل بودجه هیچ فراخوانی AJAXی نداره. محتوای پروژه‌های زیرمجموعه (PBA) از قبل توسط ویو BudgetItemsForOrgPeriodAPIView واکشی شده و مستقیماً در HTML این تمپلیت رندر می‌شه.
سطرهای تخصیص پروژه (pba): هر سطر <tr> برای تخصیص پروژه (درون بخش پروژه‌ها) (مثل data-ajax-load-url="{% url 'reports:api_tankhahs_for_pba' pba_pk=pba.pba_pk %}")، API APITankhahsForPBAView رو برای بارگذاری تنخواه‌های مربوط به اون تخصیص پروژه فراخوانی می‌کنه.

"""
class BudgetItemsForOrgPeriodAPIView(LoginRequiredMixin, View):
    """
    API View برای دریافت جزئیات BudgetAllocations (سرفصل‌ها) و
    ProjectBudgetAllocations مرتبط برای یک سازمان و دوره بودجه خاص.
    این ویو برای پاسخ به درخواست‌های AJAX از گزارش جامع استفاده می‌شود.
    """

    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request received: Get budget items for Period PK={period_pk}, Org PK={org_pk}")
        try:
            # واکشی دوره بودجه و سازمان با بررسی فعال بودن
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.debug(f"Found active BudgetPeriod: {budget_period.name} and Organization: {organization.name}")

            # ۱. واکشی BudgetAllocation ها (سرفصل‌ها) برای این سازمان و دوره
            budget_allocations_qs = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True  # فقط تخصیص‌های فعال
            ).select_related(  'budget_item',  # برای نام و کد سرفصل
                # 'project' # اگر BudgetAllocation مستقیماً به یک پروژه کلی هم می‌تواند لینک شود
            ).prefetch_related(Prefetch(
                    'project_allocations',  # related_name از BudgetAllocation به ProjectBudgetAllocation
                    queryset=ProjectBudgetAllocation.objects.filter(is_active=True)
                    .select_related('project', 'subproject')  # برای نام پروژه و زیرپروژه
                    .order_by('project__name', 'subproject__name'),
                    to_attr='active_project_allocations_list'  # نامی برای دسترسی در حلقه
                )
            ).order_by('budget_item__name')  # مرتب‌سازی بر اساس نام سرفصل

            if not budget_allocations_qs.exists():
                logger.info(
                    f"No active BudgetAllocations found for Org '{organization.name}' in Period '{budget_period.name}'.")
                # می‌توانید یک رشته HTML یا JSON خالی با پیام مناسب برگردانید
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                                {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            # ۲. آماده‌سازی داده‌ها برای ارسال
            response_data_allocations = []
            for ba_instance in budget_allocations_qs:
                # محاسبه مصرف و مانده برای این BudgetAllocation خاص
                from reports.views import calculate_total_consumed_on_budget_allocation
                from reports.views import get_budget_allocation_remaining_amount
                consumed_on_this_ba = calculate_total_consumed_on_budget_allocation(ba_instance,use_cache=False)  # False برای داده بروز در AJAX
                remaining_on_this_ba = get_budget_allocation_remaining_amount(ba_instance, use_cache=False)

                # آماده‌سازی لیست پروژه‌های تخصیص یافته از این BudgetAllocation (سرفصل)
                projects_under_this_ba = []
                # getattr برای دسترسی امن به to_attr از prefetch
                for pba in getattr(ba_instance, 'active_project_allocations_list', []):
                    project_name_display = pba.project.name if pba.project else _("پروژه نامشخص")
                    if pba.subproject:
                        project_name_display += f" / {pba.subproject.name}"

                    # مانده کلی پروژه (از تمام منابع) - این می‌تواند اختیاری باشد
                    project_overall_remaining_str = "-"
                    if pba.project:
                        try:
                            # این تابع ممکن است کوئری اضافی بزند، برای تعداد زیاد باید بهینه شود
                            from budgets.budget_calculations import get_project_remaining_budget
                            project_overall_remaining = get_project_remaining_budget(pba.project)
                            project_overall_remaining_str = f"{project_overall_remaining:,.0f}"
                        except Exception as e_proj_rem:
                            logger.warning(
                                f"Could not calculate overall remaining for project {pba.project.pk}: {e_proj_rem}")

                    projects_under_this_ba.append({
                        'pba_pk': pba.pk,
                        'project_name_display': project_name_display,
                        'allocated_to_pba': pba.allocated_amount,  # مبلغ تخصیص یافته به این پروژه/زیرپروژه از این سرفصل
                        'allocated_to_pba_formatted': f"{pba.allocated_amount:,.0f}",
                        'pba_detail_url': reverse('project_budget_allocation_detail', kwargs={'pk': pba.pk}),
                        'project_overall_remaining_formatted': project_overall_remaining_str,  # مانده کلی پروژه
                    })

                response_data_allocations.append({
                    'ba_pk': ba_instance.pk,
                    'budget_item_name': ba_instance.budget_item.name if ba_instance.budget_item else _("سرفصل نامشخص"),
                    'budget_item_code': ba_instance.budget_item.code if ba_instance.budget_item else "-",
                    'ba_allocated_amount': ba_instance.allocated_amount,
                    'ba_allocated_amount_formatted': f"{ba_instance.allocated_amount:,.0f}",
                    'ba_consumed_amount_formatted': f"{consumed_on_this_ba:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining_on_this_ba:,.0f}",
                    'ba_utilization_percentage': (
                                consumed_on_this_ba / ba_instance.allocated_amount * 100) if ba_instance.allocated_amount > 0 else 0,
                    'assigned_projects_details': projects_under_this_ba,  # لیست پروژه‌های زیرمجموعه این سرفصل
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': ba_instance.pk}),
                    # لینک به گزارش کامل این BudgetAllocation
                    # 'add_pba_url': reverse('project_budget_allocation') + f"?organization_id={ba_instance.pk}"
                    # لینک برای ایجاد PBA جدید: نیاز به organization_id دارد
                    'add_pba_url': reverse('project_budget_allocation', kwargs={'organization_id': organization.pk}) + f"?budget_allocation_id={ba_instance.pk}"
                    # لینک برای ایجاد PBA جدید
                })

            logger.info(
                f"Successfully prepared {len(response_data_allocations)} BudgetAllocation details for Org '{organization.name}' in Period '{budget_period.name}'.")

            # رندر کردن تمپلیت جزئی برای محتوای HTML
            html_content = render_to_string(
                # 'reports/partials/_budget_items_for_org_ajax.html',  # تمپلیت partial جدید
                'reports/partials/_budget_items_for_org_ajax.html',  # تمپلیت partial جدید
                {
                    'budget_allocations_data': response_data_allocations,  # نام متغیر برای تمپلیت partial
                    'parent_period_pk': period_pk,  # برای id های منحصر به فرد آکاردئون
                    'parent_org_pk': org_pk
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})

        except Http404 as e:  # اگر دوره یا سازمان یافت نشد
            logger.warning(f"API request failed: {e} (Period PK={period_pk}, Org PK={org_pk})")
            return JsonResponse({'html_content': f'<p class="text-danger text-center small py-3"><em>{e}</em></p>',
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error for budget items (Period PK={period_pk}, Org PK={org_pk}): {e}", exc_info=True)
            return JsonResponse(
                {'error': _("خطا در پردازش درخواست. لطفاً با پشتیبانی تماس بگیرید."), 'status': 'error'}, status=500)
class YourOrgPeriodAllocationsListView(LoginRequiredMixin, ListView):
    model = BudgetAllocation
    template_name = 'reports/org_period_allocations_list.html'  # یک تمپلیت جدید برای این لیست
    context_object_name = 'budget_allocations'
    paginate_by = 15

    # permission_codenames = ['budgets.view_budgetallocation'] # یا پرمیشن مناسب

    def get_queryset(self):
        period_pk = self.kwargs.get('period_pk')
        org_pk = self.kwargs.get('org_pk')

        if not period_pk or not org_pk:
            logger.warning("OrgPeriodAllocationsList: Missing period_pk or org_pk in URL.")
            return BudgetAllocation.objects.none()

        self.budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True)
        self.organization = get_object_or_404(Organization, pk=org_pk)

        logger.info(
            f"Fetching BudgetAllocations for Org '{self.organization.name}' in Period '{self.budget_period.name}'")

        queryset = BudgetAllocation.objects.filter(
            budget_period=self.budget_period,
            organization=self.organization,
            is_active=True
        ).select_related(
            'budget_item', 'project', 'created_by'  # و سایر فیلدهای لازم
        ).order_by('budget_item__name')

        # می‌توانید فیلترهای جستجو و ... را هم اینجا اضافه کنید
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['report_title'] = _("لیست سرفصل‌های بودجه تخصیص یافته")
        context['budget_period_instance'] = getattr(self, 'budget_period', None)  # برای نمایش در هدر
        context['organization_instance'] = getattr(self, 'organization', None)  # برای نمایش در هدر
        logger.debug(
            f"Context for OrgPeriodAllocationsList: Period={context['budget_period_instance']}, Org={context['organization_instance']}")
        return context
#--
class OrganizationAllocationsAPIView(LoginRequiredMixin, View):
    def get(self, request, period_pk, *args, **kwargs):
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)

            organization_summaries = []
            org_alloc_queryset = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            for org_summary in org_alloc_queryset:
                # محاسبه مصرف (این بخش همچنان نیاز به بهینه‌سازی دارد اگر تعداد زیاد است)
                consumed = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed = consumed - returned

                organization_summaries.append({
                    'id': org_summary['organization__id'],
                    'name': org_summary['organization__name'],
                    'code': org_summary['organization__code'],
                    'total_allocated_formatted': f"{org_summary['total_allocated_to_org']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{org_summary['total_allocated_to_org'] - net_consumed:,.0f}",
                    'num_budget_items': org_summary['num_budget_items'],
                    'utilization_percentage': (net_consumed / org_summary['total_allocated_to_org'] * 100) if
                    org_summary['total_allocated_to_org'] > 0 else 0,
                    # URL برای بارگذاری سرفصل‌های این سازمان
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_summary['organization__id']})
                })
            return JsonResponse({'organizations': organization_summaries, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("دوره بودجه یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (OrgAllocations) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش."), 'status': 'error'}, status=500)
##############################

class ComprehensiveBudgetReportView(LoginRequiredMixin, ListView):
    model = BudgetPeriod
    # template_name = 'reports/v2/comprehensive_report_main.html'  # مسیر صحیح تمپلیت
    template_name = 'reports/v2/comprehensive_report_main.html'  # مسیر صحیح تمپلیت
    context_object_name = 'budget_periods_report_data_from_context'

    # paginate_by = 5 # اگر تعداد دوره‌های بودجه زیاد است، فعال کنید

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Executing get_queryset...")
        queryset = BudgetPeriod.objects.filter(
            is_active=True,
            is_completed=False
        ).select_related(
            'organization',
        ).order_by('-start_date', 'name')

        search_query = self.request.GET.get('search_period', None)
        if search_query:
            logger.debug(f"Applying search filter for period name: '{search_query}'")
            queryset = queryset.filter(name__icontains=search_query)

        count = queryset.count()
        logger.info(f"[{self.__class__.__name__}] - get_queryset finished. Found {count} BudgetPeriods.")
        if count == 0:
            logger.warning("No active and non-completed BudgetPeriods found based on current filters.")
        return queryset

    def _get_processed_report_data(self, budget_periods_list):
        processed_periods_output = []
        if not budget_periods_list:
            logger.warning("[_get_processed_report_data] - Received an empty list of budget periods.")
            return processed_periods_output

        for period_instance in budget_periods_list:
            logger.debug(f"Processing data for BudgetPeriod: {period_instance.name} (PK: {period_instance.pk})")

            allocations_in_period_qs = BudgetAllocation.objects.filter(budget_period=period_instance, is_active=True)
            total_allocated_from_this_period = allocations_in_period_qs.aggregate(
                total=Coalesce(Sum('allocated_amount'), Decimal('0'), output_field=DecimalField())
            )['total']

            all_transactions_in_period_qs = BudgetTransaction.objects.filter(allocation__budget_period=period_instance)
            consumed_in_period = all_transactions_in_period_qs.filter(transaction_type='CONSUMPTION').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            returned_in_period = all_transactions_in_period_qs.filter(transaction_type='RETURN').aggregate(
                total=Coalesce(Sum('amount'), Decimal('0'), output_field=DecimalField())
            )['total']
            net_consumed_from_this_period = consumed_in_period - returned_in_period

            remaining_of_period_vs_total_amount = period_instance.total_amount - net_consumed_from_this_period
            utilization_of_period_vs_total_amount = (net_consumed_from_this_period / period_instance.total_amount * 100) \
                if period_instance.total_amount > 0 else Decimal('0')

            organization_summaries = []
            org_alloc_queryset_summary = BudgetAllocation.objects.filter(
                budget_period=period_instance, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org_from_this_period=Sum('allocated_amount'),
                num_budget_items_for_org=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            logger.debug(
                f"Period {period_instance.pk} - Found {org_alloc_queryset_summary.count()} distinct organizations with allocations.")

            for org_summary_data in org_alloc_queryset_summary:
                org_id = org_summary_data['organization__id']
                logger.debug(
                    f"  Processing org summary for Org ID: {org_id} ({org_summary_data['organization__name']})")

                # **بهینه‌سازی مصرف سازمان (مهم برای پرفورمنس):**
                # این بخش را می‌توان با یک Subquery در annotate اصلی org_alloc_queryset_summary انجام داد.
                # برای این مثال، همچنان جدا محاسبه می‌کنیم.
                consumed_by_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_id,
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned_to_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_id,
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed_by_this_org_val = consumed_by_this_org - returned_to_this_org

                remaining_for_this_org_val = org_summary_data[
                                                 'total_allocated_to_org_from_this_period'] - net_consumed_by_this_org_val

                ajax_detail_url = reverse('api_budget_items_for_org_period',  # نام URL از reports.urls
                                          kwargs={'period_pk': period_instance.pk, 'org_pk': org_id})

                first_ba_instance = BudgetAllocation.objects.filter(
                    budget_period=period_instance,
                    organization_id=org_id,
                    is_active=True
                ).select_related('budget_item').order_by('budget_item__name',
                                                         'pk').first()  # ترتیب برای انتخاب اولین سرفصل

                organization_summaries.append({
                    'id': org_id,
                    'name': org_summary_data['organization__name'],
                    'code': org_summary_data['organization__code'],
                    'total_allocated': org_summary_data['total_allocated_to_org_from_this_period'],
                    'net_consumed': net_consumed_by_this_org_val,
                    'remaining': remaining_for_this_org_val,
                    'num_budget_items': org_summary_data['num_budget_items_for_org'],
                    'utilization_percentage': (net_consumed_by_this_org_val / org_summary_data[
                        'total_allocated_to_org_from_this_period'] * 100) \
                        if org_summary_data['total_allocated_to_org_from_this_period'] > 0 else Decimal('0'),
                    'ajax_load_url_for_budget_items': ajax_detail_url,  # نام را تغییر دادم برای وضوح
                    'first_ba_instance_pk_for_report_link': first_ba_instance.pk if first_ba_instance else None,
                })

            processed_periods_output.append({
                'period': period_instance,
                'summary': {
                    'total_budget': period_instance.total_amount,
                    'total_allocated_from_period': total_allocated_from_this_period,
                    'net_consumed_from_period': net_consumed_from_this_period,
                    'remaining_vs_consumed': remaining_of_period_vs_total_amount,
                    'utilization_percentage': utilization_of_period_vs_total_amount,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period_instance.start_date).strftime(
                        '%Y/%m/%d') if period_instance.start_date else "-",
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period_instance.end_date).strftime(
                        '%Y/%m/%d') if period_instance.end_date else "-",
                },
                'organization_summaries': organization_summaries,  # نام را برای سازگاری با تمپلیت قبلی تغییر دادم
            })
        logger.info(
            f"[_get_processed_report_data] - Finished processing. Returning {len(processed_periods_output)} items.")
        return processed_periods_output

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        logger.info(
            f"[{self.__class__.__name__}] - Starting get_context_data. Initial object_list count: {len(context.get('object_list', []))}")

        report_data = self._get_processed_report_data(context.get('object_list', []))

        context[self.context_object_name] = report_data
        context['report_main_title'] = _("گزارش جامع بودجه و تخصیص‌ها")

        # برای فرم فیلتر در بالای صفحه
        context['current_search_period'] = self.request.GET.get('search_period', '')

        logger.info(
            f"[{self.__class__.__name__}] - get_context_data finished. Final context_object_name count: {len(report_data)}")
        if not report_data and not self.request.GET.get('search_period'):  # فقط اگر جستجویی هم در کار نبوده
            messages.info(self.request, _("هیچ دوره بودجه فعالی برای نمایش یافت نشد."))
        elif not report_data and self.request.GET.get('search_period'):
            messages.info(self.request, _("هیچ دوره بودجه‌ای با معیارهای جستجوی شما یافت نشد."))

        return context

    def generate_excel_report(self, report_data_for_export):
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library is not installed. Cannot generate Excel report.")
            return None

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="comprehensive_budget_report.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True

        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFFFF")  # فونت سفید
        header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")  # آبی تیره
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_align = Alignment(horizontal='right', vertical='center')
        number_format_currency = '#,##0_-"ریال";-#,##0_-"ریال";0_-"ریال";@'  # فرمت ارز فارسی
        number_format_percentage = '0.0%'
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل دوره (ریال)"),
            _("تخصیص از دوره (ریال)"), _("مصرف از دوره (ریال)"), _("مانده دوره (ریال)"), _("% مصرف دوره"),
            _("سازمان تخصیص گیرنده"), _("کد"), _("تخصیص به سازمان (ریال)"), _("مصرف سازمان (ریال)"),
            _("مانده سازمان (ریال)"), _("% مصرف سازمان")
        ]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font;
            cell.fill = header_fill;
            cell.alignment = header_align;
            cell.border = thin_border
            sheet.column_dimensions[get_column_letter(col_num)].width = 20 if col_num > 9 else 25

        row_num = 2
        for period_data in report_data_for_export:
            p_instance = period_data['period_instance']  # تغییر نام متغیر برای سازگاری با get_report_data_for_html
            p_summary = period_data['summary']

            start_row_for_period_merge = row_num

            # اگر هیچ سازمانی برای این دوره بودجه تخصیص نگرفته است، فقط اطلاعات دوره را چاپ کن
            if not period_data['organization_summaries']:
                sheet.cell(row=row_num, column=1, value=p_instance.name).border = thin_border
                sheet.cell(row=row_num, column=2, value=p_instance.organization.name).border = thin_border
                sheet.cell(row=row_num, column=3, value=p_summary['start_date_jalali']).border = thin_border
                sheet.cell(row=row_num, column=4, value=p_summary['end_date_jalali']).border = thin_border
                sheet.cell(row=row_num, column=5,
                           value=p_summary['total_budget']).number_format = number_format_currency
                sheet.cell(row=row_num, column=6,
                           value=p_summary['total_allocated_from_period']).number_format = number_format_currency
                sheet.cell(row=row_num, column=7,
                           value=p_summary['net_consumed_from_period']).number_format = number_format_currency
                sheet.cell(row=row_num, column=8,
                           value=p_summary['remaining_vs_consumed']).number_format = number_format_currency
                sheet.cell(row=row_num, column=9,
                           value=p_summary['utilization_percentage'] / 100).number_format = number_format_percentage
                for c_idx in range(1, 10): sheet.cell(row=row_num, column=c_idx).border = thin_border
                row_num += 1
            else:
                for org_sum_data in period_data['organization_summaries']:
                    # اطلاعات دوره بودجه
                    sheet.cell(row=row_num, column=1, value=p_instance.name)
                    sheet.cell(row=row_num, column=2, value=p_instance.organization.name)
                    sheet.cell(row=row_num, column=3, value=p_summary['start_date_jalali'])
                    sheet.cell(row=row_num, column=4, value=p_summary['end_date_jalali'])
                    sheet.cell(row=row_num, column=5,
                               value=p_summary['total_budget']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=6,
                               value=p_summary['total_allocated_from_period']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=7,
                               value=p_summary['net_consumed_from_period']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=8,
                               value=p_summary['remaining_vs_consumed']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=9,
                               value=p_summary['utilization_percentage'] / 100).number_format = number_format_percentage

                    # اطلاعات سازمان
                    sheet.cell(row=row_num, column=10, value=org_sum_data['name'])
                    sheet.cell(row=row_num, column=11, value=org_sum_data['code'])
                    sheet.cell(row=row_num, column=12,
                               value=org_sum_data['total_allocated']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=13,
                               value=org_sum_data['net_consumed']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=14,
                               value=org_sum_data['remaining']).number_format = number_format_currency
                    sheet.cell(row=row_num, column=15, value=org_sum_data[
                                                                 'utilization_percentage'] / 100).number_format = number_format_percentage

                    for c_idx in range(1, 16): sheet.cell(row=row_num, column=c_idx).border = thin_border
                    row_num += 1

                # مرج کردن سلول‌های دوره بودجه اگر بیش از یک سازمان دارد
                if len(period_data['organization_summaries']) > 1:
                    for col_idx_merge in range(1, 10):  # ستون‌های مربوط به دوره
                        sheet.merge_cells(start_row=start_row_for_period_merge, start_column=col_idx_merge,
                                          end_row=row_num - 1, end_column=col_idx_merge)
                        sheet.cell(row=start_row_for_period_merge, column=col_idx_merge).alignment = Alignment(
                            vertical='top', horizontal='right')
        try:
            workbook.save(response)
            return response
        except Exception as e_excel:
            logger.error(f"Error saving Excel workbook: {e_excel}", exc_info=True)
            return None

    def generate_pdf_report(self, report_data_for_export, report_title):
        if not WEASYPRINT_AVAILABLE:
            logger.error("WeasyPrint library is not installed. Cannot generate PDF report.")
            return None

        html_string = render_to_string(
            'reports/pdf/comprehensive_budget_report_pdf.html',
            {
                'budget_periods_report_data_for_pdf': report_data_for_export,  # نام جدید
                'report_main_title': report_title
            }
        )
        # استایل ساده برای PDF
        pdf_css_string_internal = """
            @page { size: A3 landscape; margin: 1.2cm; @bottom-center { content: "صفحه " counter(page) " از " counter(pages); font-size: 7pt; color: #555;} }
            body { font-family: "Tahoma", "B Nazanin", "DejaVu Sans", sans-serif; direction: rtl; font-size: 8pt; line-height: 1.4;}
            table { width: 100%; border-collapse: collapse; margin-bottom: 10px; page-break-inside: avoid; }
            th, td { border: 1px solid #999; padding: 4px; text-align: right; word-wrap: break-word; font-size: 7.5pt; }
            th { background-color: #f0f0f0; font-weight: bold; }
            h1 { font-size: 16pt; text-align:center; margin-bottom: 5px;} 
            h2 { font-size: 12pt; text-align:right; margin-top:15px; margin-bottom: 5px; padding-bottom:3px; border-bottom: 1px solid #ccc;} 
            h3 { font-size: 10pt; text-align:right; margin-top:10px; margin-bottom: 4px;}
            .period-summary-pdf { background-color: #f9f9f9; padding: 6px; margin-bottom:8px; border: 1px solid #dedede; font-size: 7.5pt;}
            .text-danger-pdf { color: #c00; } .text-success-pdf { color: #060; }
            .small-text { font-size: 7pt; color: #444; }
        """
        try:
            html_doc = HTML(string=html_string, base_url=self.request.build_absolute_uri())
            pdf_file = html_doc.write_pdf(stylesheets=[CSS(string=pdf_css_string_internal)])

            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="Comprehensive_Budget_Report.pdf"'
            return response
        except Exception as e_pdf:
            logger.error(f"Error generating PDF with WeasyPrint: {e_pdf}", exc_info=True)
            return None

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        # داده‌های گزارش از context که توسط get_context_data پر شده، استخراج می‌شود
        # context_object_name ما 'budget_periods_report_data_from_context' است
        report_data_from_context = context.get('budget_periods_report_data_from_context', [])
        report_title_from_context = context.get('report_main_title', _("گزارش جامع بودجه"))

        logger.debug(
            f"Render_to_response called. Output format: {output_format}. Report data items: {len(report_data_from_context)}")

        if output_format == 'pdf':
            pdf_response = self.generate_pdf_report(report_data_from_context, report_title_from_context)
            if pdf_response:
                return pdf_response
            else:
                messages.error(self.request,
                               _("خطا در تولید خروجی PDF. لطفاً از نصب بودن WeasyPrint و فونت‌های فارسی مناسب اطمینان حاصل کنید."))

        elif output_format == 'excel':
            excel_response = self.generate_excel_report(report_data_from_context)
            if excel_response:
                return excel_response
            else:
                messages.error(self.request,
                               _("خطا در تولید خروجی اکسل. لطفاً از نصب بودن openpyxl اطمینان حاصل کنید."))

        # اگر فرمت HTML یا خطایی در خروجی دیگر رخ داده، تمپلیت اصلی را رندر کن
        # ListView به طور خودکار object_list را در context قرار می‌دهد.
        # ما در get_context_data، مقدار context[self.context_object_name] را با داده‌های پردازش شده خودمان جایگزین کردیم.
        # برای اینکه paginate_by کار کند، ListView به context['object_list'] نیاز دارد که همان get_queryset() اولیه باشد.
        # اما چون ما همه داده‌ها را در _get_processed_report_data پردازش می‌کنیم،
        # صفحه‌بندی باید روی نتیجه نهایی (report_data) اعمال شود که پیچیده‌تر است.
        # برای این مثال، paginate_by را غیرفعال نگه می‌داریم.
        # اگر صفحه‌بندی در سطح دوره‌های بودجه مهم است، باید _get_processed_report_data فقط برای آیتم‌های صفحه فعلی اجرا شود.
        return super().render_to_response(context, **response_kwargs)
#--------- API
class OLD_APIOrganizationAllocationsView(LoginRequiredMixin, View):
    def get(self, request, period_pk, *args, **kwargs):
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)

            # کوئری بهینه شده برای خلاصه سازمان‌ها
            org_summaries_qs = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                # محاسبه مصرف با Subquery برای بهینگی (نیاز به Django 2.0+)
                # یا محاسبه جداگانه در پایتون اگر Subquery پیچیده است
                # مثال با محاسبه جداگانه (برای سادگی، اما برای پرفورمنس باید بهینه شود):
            ).order_by('organization__name')

            organizations_data = []
            for org_sum in org_summaries_qs:
                # این محاسبه مصرف باید بهینه شود (مثلا با یک annotate جداگانه روی queryset اصلی)
                consumed_val = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_sum['organization__id'],
                    transaction_type='CONSUMPTION').aggregate(s=Coalesce(Sum('amount'), Decimal(0)))['s']
                returned_val = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_sum['organization__id'],
                    transaction_type='RETURN').aggregate(s=Coalesce(Sum('amount'), Decimal(0)))['s']
                net_consumed_val = consumed_val - returned_val

                organizations_data.append({
                    'id': org_sum['organization__id'],
                    'name': org_sum['organization__name'],
                    'code': org_sum['organization__code'],
                    'total_allocated_formatted': f"{org_sum['total_allocated']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed_val:,.0f}",
                    'remaining_formatted': f"{org_sum['total_allocated'] - net_consumed_val:,.0f}",
                    'num_budget_items': org_sum['num_budget_items'],
                    'utilization_percentage': (net_consumed_val / org_sum['total_allocated'] * 100) if org_sum[
                                                                                                           'total_allocated'] > 0 else 0,
                    'budget_items_ajax_url': reverse('api_budget_allocations_for_org',
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_sum['organization__id']})
                })

            # رندر کردن تمپلیت جزئی برای این سطح
            html_content = render_to_string('reports/partials/_ajax_level_organizations.html',
                                            {'organizations': organizations_data, 'parent_period_pk': period_pk})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("دوره بودجه یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (OrgAllocations) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش سازمان‌ها."), 'status': 'error'}, status=500)
class APIBudgetAllocationsForOrgView(LoginRequiredMixin, View):
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk)
            organization = get_object_or_404(Organization, pk=org_pk)

            budget_allocations_qs = BudgetAllocation.objects.filter(
                budget_period=period, organization=organization, is_active=True
            ).select_related('budget_item').order_by('budget_item__name')

            allocations_data = []
            for ba in budget_allocations_qs:
                from reports.views import calculate_total_consumed_on_budget_allocation,get_budget_allocation_remaining_amount
                consumed = calculate_total_consumed_on_budget_allocation(ba, use_cache=False)
                remaining = get_budget_allocation_remaining_amount(ba, use_cache=False)
                allocations_data.append({
                    'id': ba.pk,
                    'item_name': ba.budget_item.name if ba.budget_item else _("سرفصل نامشخص"),
                    'item_code': ba.budget_item.code if ba.budget_item else "-",
                    'allocated_formatted': f"{ba.allocated_amount:,.0f}",
                    'consumed_formatted': f"{consumed:,.0f}",
                    'remaining_formatted': f"{remaining:,.0f}",
                    'utilization_percentage': (consumed / ba.allocated_amount * 100) if ba.allocated_amount > 0 else 0,
                    'project_allocations_ajax_url': reverse('api_project_allocations_for_ba',
                                                            kwargs={'ba_pk': ba.pk}),
                    'report_url': reverse('budget_allocation_report', kwargs={'pk': ba.pk}),
                    # لینک به گزارش جزئیات BudgetAllocation
                    'add_pba_url': reverse('project_budget_allocation') + f"?organization_id={ba.pk}"
                })

            html_content = render_to_string('reports/partials/_ajax_level_budget_allocations.html',
                                            {'budget_allocations': allocations_data, 'parent_org_pk': org_pk,
                                             'parent_period_pk': period_pk})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("دوره یا سازمان یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (BudgetAllocations) for Period {period_pk}, Org {org_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش سرفصل‌های بودجه."), 'status': 'error'}, status=500)
class OLD_APIProjectAllocationsForBAView(LoginRequiredMixin, View):
    def get(self, request, ba_pk, *args, **kwargs):
        try:
            budget_allocation = get_object_or_404(BudgetAllocation.objects.select_related('budget_item'), pk=ba_pk)

            project_allocs_qs = ProjectBudgetAllocation.objects.filter(
                budget_allocation=budget_allocation, is_active=True
            ).select_related('project', 'subproject').order_by('project__name', 'subproject__name')

            project_allocations_data = []
            for pba in project_allocs_qs:
                target_name = pba.project.name if pba.project else ""
                if pba.subproject:
                    target_name += f" / {pba.subproject.name}"

                # محاسبه مصرف و مانده برای این PBA (باید دقیق‌تر شود)
                # این تابع get_project_remaining_budget ممکن است کل پروژه را در نظر بگیرد
                # شما به مصرف از *این* PBA خاص یا مصرف کلی پروژه/زیرپروژه نیاز دارید
                target_remaining = "-"
                if pba.subproject:
                    # target_remaining = get_subproject_remaining_budget(pba.subproject)
                    pass  # برای سادگی فعلا خالی
                elif pba.project:
                    # target_remaining = get_project_remaining_budget(pba.project)
                    pass  # برای سادگی فعلا خالی

                project_allocations_data.append({
                    'id': pba.pk,
                    'target_name': target_name,
                    'allocated_formatted': f"{pba.allocated_amount:,.0f}",
                    'remaining_formatted': f"{target_remaining}",  # نیاز به محاسبه دقیق
                    'detail_url': reverse('project_budget_allocation_detail', kwargs={'pk': pba.pk}),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_pba', kwargs={'pba_pk': pba.pk}),
                    'add_tankhah_url': reverse('tankhah:tankhah_create') + f"?project_budget_allocation_id={pba.pk}"
                    # یا pba.project.id و ...
                })

            html_content = render_to_string('reports/partials/_ajax_level_project_allocations.html',
                                            {'project_allocations': project_allocations_data, 'parent_ba_pk': ba_pk,
                                             'budget_item_name': budget_allocation.budget_item.name})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("تخصیص بودجه یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (ProjectAllocations) for BA PK={ba_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش تخصیص‌های پروژه."), 'status': 'error'}, status=500)
#----------------
# --- ComprehensiveBudgetReportView (ویو اصلی) ---
# (کد این ویو مانند پاسخ قبلی است، فقط مطمئن شوید که از نام‌های URL صحیح در
# بخش `ajax_load_organizations_url` در `_get_processed_report_data` استفاده می‌کند
# یعنی: reverse('api_organization_allocations_for_period', ...))
# و همچنین در بخش generate_excel_report و generate_pdf_report،
# اگر می‌خواهید گزارش کامل سلسله مراتبی باشد، باید یک متد جداگانه برای واکشی *تمام* داده‌ها بنویسید.

# --- APIOrganizationAllocationsView (برای بارگذاری سازمان‌های یک دوره) ---
# (کد این ویو مانند پاسخ قبلی است، فقط مطمئن شوید از نام URL صحیح برای
# `budget_items_ajax_url` استفاده می‌کند:
# reverse('api_budget_allocations_for_org', ...))

class APIOrganizationAllocationsView(LoginRequiredMixin, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API request: Get organizations for Period PK={period_pk}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            org_summaries_qs = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            organizations_data = []
            for org_sum in org_summaries_qs:
                # این محاسبه مصرف باید بهینه شود (مثلا با یک annotate جداگانه روی queryset اصلی)
                consumed_val = BudgetTransaction.objects.filter(allocation__budget_period=period,
                                                                allocation__organization_id=org_sum['organization__id'],
                                                                transaction_type='CONSUMPTION').aggregate(
                    s=Coalesce(Sum('amount'), Decimal(0)))['s']
                returned_val = BudgetTransaction.objects.filter(allocation__budget_period=period,
                                                                allocation__organization_id=org_sum['organization__id'],
                                                                transaction_type='RETURN').aggregate(
                    s=Coalesce(Sum('amount'), Decimal(0)))['s']
                net_consumed_val = consumed_val - returned_val

                organizations_data.append({
                    'id': org_sum['organization__id'],
                    'name': org_sum['organization__name'],
                    'code': org_sum['organization__code'],
                    'total_allocated_formatted': f"{org_sum['total_allocated']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed_val:,.0f}",
                    'remaining_formatted': f"{org_sum['total_allocated'] - net_consumed_val:,.0f}",
                    'num_budget_items': org_sum['num_budget_items'],
                    'utilization_percentage': (net_consumed_val / org_sum['total_allocated'] * 100) if org_sum[
                                                                                                           'total_allocated'] > 0 else 0,
                    'budget_items_ajax_url': reverse('api_budget_allocations_for_org',  # نام صحیح URL
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_sum['organization__id']})
                })
            html_content = render_to_string('reports/partials/_ajax_level_organizations.html',
                                            {'organizations': organizations_data, 'parent_period_pk': period_pk})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'html_content': render_to_string('reports/partials/_no_organizations_found.html',
                                                                  {'message': _("دوره بودجه یافت نشد.")}),
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (OrgAllocations) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse({'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                                  {'error_message': _("خطا در پردازش سازمان‌ها.")}),
                                 'status': 'error'}, status=500)
class erro__APIBudgetAllocationsForOrgView(LoginRequiredMixin, View):
    """
    API برای دریافت BudgetAllocations (سرفصل‌ها) برای یک سازمان و دوره خاص.
    """

    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request: Get BudgetAllocations for Period PK={period_pk}, Org PK={org_pk}")
        try:
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)

            budget_allocations_qs = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').order_by('budget_item__name')

            if not budget_allocations_qs.exists():
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                                {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            allocations_data = []
            for ba_instance in budget_allocations_qs:
                from reports.views import calculate_total_consumed_on_budget_allocation
                consumed_on_this_ba = calculate_total_consumed_on_budget_allocation(ba_instance, use_cache=False)
                from reports.views import get_budget_allocation_remaining_amount
                remaining_on_this_ba = get_budget_allocation_remaining_amount(ba_instance, use_cache=False)

                allocations_data.append({
                    'ba_pk': ba_instance.pk,
                    'budget_item_name': ba_instance.budget_item.name if ba_instance.budget_item else _("سرفصل نامشخص"),
                    'budget_item_code': ba_instance.budget_item.code if ba_instance.budget_item else "-",
                    'ba_allocated_amount_formatted': f"{ba_instance.allocated_amount:,.0f}",
                    'ba_consumed_amount_formatted': f"{consumed_on_this_ba:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining_on_this_ba:,.0f}",
                    'ba_utilization_percentage': (
                                consumed_on_this_ba / ba_instance.allocated_amount * 100) if ba_instance.allocated_amount > 0 else 0,
                    # URL برای بارگذاری ProjectBudgetAllocation های این سرفصل
                    'project_allocations_ajax_url': reverse('api_project_allocations_for_ba',
                                                            kwargs={'ba_pk': ba_instance.pk}),
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': ba_instance.pk}),
                    'add_pba_url': reverse(
                        'project_budget_allocation_create') + f"?budget_allocation_id={ba_instance.pk}"
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {
                    'budget_allocations_list_data': allocations_data,  # نام را برای وضوح در تمپلیت تغییر دادم
                    'parent_period_pk': period_pk,
                    'parent_org_pk': org_pk,
                    'parent_budget_period_name': budget_period.name,
                    'parent_organization_name': organization.name
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'html_content': render_to_string('reports/partials/_no_data_found_ajax.html',
                                                                  {'message': _("دوره یا سازمان یافت نشد.")}),
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (BudgetAllocations) for Period {period_pk}, Org {org_pk}: {e}", exc_info=True)
            return JsonResponse({'html_content': render_to_string('reports/partials/_error_ajax.html', {
                'error_message': _("خطا در پردازش سرفصل‌های بودجه.")}), 'status': 'error'}, status=500)
class APIProjectAllocationsForBAView(LoginRequiredMixin, View):
    """
    API برای دریافت ProjectBudgetAllocations برای یک BudgetAllocation (سرفصل) خاص.
    """

    def get(self, request, ba_pk, *args, **kwargs):
        logger.info(f"API request: Get ProjectBudgetAllocations for BA PK={ba_pk}")
        try:
            budget_allocation = get_object_or_404(
                BudgetAllocation.objects.select_related('budget_item', 'organization', 'budget_period'), pk=ba_pk)

            project_allocs_qs = ProjectBudgetAllocation.objects.filter(
                budget_allocation=budget_allocation, is_active=True
            ).select_related('project', 'subproject').order_by('project__name', 'subproject__name')

            if not project_allocs_qs.exists():
                html_content = render_to_string('reports/partials/_no_project_allocations_found.html',
                                                {'budget_allocation': budget_allocation})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            project_allocations_data = []
            for pba in project_allocs_qs:
                target_name = pba.project.name if pba.project else _("نامشخص")
                if pba.subproject:
                    target_name += f" / {pba.subproject.name}"

                # محاسبه مانده برای این PBA یا پروژه/زیرپروژه کلی
                # برای این مثال، مانده خود PBA را نمایش می‌دهیم (نیاز به تابع یا متد get_remaining_amount در PBA)
                # یا مانده کلی پروژه/زیرپروژه
                pba_remaining_val = pba.get_remaining_amount() if hasattr(pba,
                                                                          'get_remaining_amount') else pba.allocated_amount  # مثال ساده شده
                # یا
                # project_overall_remaining_val = "-"
                # if pba.subproject: project_overall_remaining_val = get_subproject_remaining_budget(pba.subproject)
                # elif pba.project: project_overall_remaining_val = get_project_remaining_budget(pba.project)

                project_allocations_data.append({
                    'pba_pk': pba.pk,
                    'target_name_display': target_name,
                    'allocated_to_pba_formatted': f"{pba.allocated_amount:,.0f}",
                    'pba_remaining_formatted': f"{pba_remaining_val:,.0f}",  # این باید محاسبه شود
                    'pba_detail_url': reverse('project_budget_allocation_detail', kwargs={'pk': pba.pk}),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_pba', kwargs={'pba_pk': pba.pk}),
                    'add_tankhah_url': reverse('tankhah:tankhah_create') + f"?project_budget_allocation_id={pba.pk}"
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_project_allocations.html',
                {
                    'project_allocations_list_data': project_allocations_data,  # نام جدید
                    'parent_ba_pk': ba_pk,
                    'parent_budget_item_name': budget_allocation.budget_item.name if budget_allocation.budget_item else "",
                    'parent_org_name': budget_allocation.organization.name,
                    'parent_period_name': budget_allocation.budget_period.name
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'html_content': render_to_string('reports/partials/_no_data_found_ajax.html',
                                                                  {'message': _("تخصیص بودجه سرفصل یافت نشد.")}),
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (ProjectAllocations) for BA PK={ba_pk}: {e}", exc_info=True)
            return JsonResponse({'html_content': render_to_string('reports/partials/_error_ajax.html', {
                'error_message': _("خطا در پردازش تخصیص‌های پروژه.")}), 'status': 'error'}, status=500)
"""
شرح تک خطی: این API، لیست تنخواه‌ها رو برای یک تخصیص پروژه مشخص برمی‌گردونه.
تمپلیت مرتبط: reports/partials/_ajax_level_tankhahs.html (این HTML رو در پاسخ JSON برمی‌گردونه).
نحوه فراخوانی API‌ها در این تمپلیت:
سطرهای تنخواه: هر سطر <tr> در این تمپلیت (مثل data-ajax-load-url="{% url 'reports:api_factors_for_tankhah' tankhah_pk=tankhah.id %}")، API APIFactorsForTankhahView رو برای بارگذاری فاکتورهای مربوط به اون تنخواه فراخوانی می‌کنه.
"""
class APITankhahsForPBAView(LoginRequiredMixin, View):
    def get(self, request, pba_pk, *args, **kwargs):
        logger.info(f"API request: Get Tankhahs for PBA PK={pba_pk}")
        try:
            pba_instance = get_object_or_404(ProjectBudgetAllocation.objects.select_related('project', 'subproject'),pk=pba_pk)

            # تنخواه‌های مرتبط با این ProjectBudgetAllocation
            from tankhah.models import Tankhah
            tankhahs_qs = Tankhah.objects.filter(project_budget_allocation=pba_instance).select_related('created_by',
                                                                                                        'current_stage').order_by(
                '-date')

            if not tankhahs_qs.exists():
                html_content = render_to_string('reports/partials/_no_tankhahs_found.html',
                                                {'parent_pba': pba_instance})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            tankhahs_data = []
            for tankhah in tankhahs_qs:
                tankhahs_data.append({
                    'id': tankhah.pk,
                    'number': tankhah.number,
                    'amount_formatted': f"{tankhah.amount:,.0f}",
                    'status_display': tankhah.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=tankhah.date).strftime(
                        '%Y/%m/%d') if tankhah.date else "-",
                    'detail_url': reverse('tankhah:tankhah_detail', kwargs={'pk': tankhah.pk}),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': tankhah.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {
                    'tankhahs_list_data': tankhahs_data,  # نام جدید
                    'parent_pba_pk': pba_pk,
                    'parent_pba_name': f"{pba_instance.project.name}{f' / {pba_instance.subproject.name}' if pba_instance.subproject else ''}"
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'html_content': render_to_string('reports/partials/_no_data_found_ajax.html',
                                                                  {'message': _("تخصیص بودجه پروژه یافت نشد.")}),
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (Tankhahs) for PBA PK={pba_pk}: {e}", exc_info=True)
            return JsonResponse({'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                                  {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                                 'status': 'error'}, status=500)
"""
شرح تک خطی: این API، لیست فاکتورها رو برای یک تنخواه مشخص برمی‌گردونه.
تمپلیت مرتبط: reports/partials/_ajax_level_factors.html (این HTML رو در پاسخ JSON برمی‌گردونه).
نحوه فراخوانی API‌ها در این تمپلیت:

سطرهای فاکتور: این آخرین سطح گزارش است و معمولاً هیچ API دیگری رو فراخوانی نمی‌کنه.
"""
class APIFactorsForTankhahView(LoginRequiredMixin, View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API request: Get Factors for Tankhah PK={tankhah_pk}")
        try:
            from tankhah.models import Tankhah
            tankhah_instance = get_object_or_404(Tankhah, pk=tankhah_pk)
            from tankhah.models import Factor
            factors_qs = Factor.objects.filter(tankhah=tankhah_instance).select_related('category',
                                                                                        'created_by').order_by('-date')

            if not factors_qs.exists():
                html_content = render_to_string('reports/partials/_no_factors_found.html',
                                                {'parent_tankhah': tankhah_instance})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            factors_data = []
            for factor in factors_qs:
                factors_data.append({
                    'id': factor.pk,
                    'number': factor.number,
                    'amount_formatted': f"{factor.amount:,.0f}",
                    'status_display': factor.get_status_display(),
                    'category_name': factor.category.name if factor.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=factor.date).strftime(
                        '%Y/%m/%d') if factor.date else "-",
                    'detail_url': reverse('tankhah:factor_detail', kwargs={'factor_pk': factor.pk}),  # یا نام URL شما
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_factors.html',
                {
                    'factors_list_data': factors_data,  # نام جدید
                    'parent_tankhah_pk': tankhah_pk,
                    'parent_tankhah_number': tankhah_instance.number
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse({'html_content': render_to_string('reports/partials/_no_data_found_ajax.html',
                                                                  {'message': _("تنخواه یافت نشد.")}),
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (Factors) for Tankhah PK={tankhah_pk}: {e}", exc_info=True)
            return JsonResponse({'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                                  {'error_message': _("خطا در پردازش فاکتورها.")}),
                                 'status': 'error'}, status=500)
# --- ویوهای API برای سطوح پایین‌تر (Tankhah) باید به همین ترتیب ایجاد شوند ---
class OLD_APITankhahsForPBAView(LoginRequiredMixin, View):
    def get(self, request, pba_pk, *args, **kwargs):
        # ۱. pba_instance = get_object_or_404(ProjectBudgetAllocation, pk=pba_pk)
        # ۲. tankhahs = Tankhah.objects.filter(project_budget_allocation=pba_instance) یا اگر لینک مستقیم به پروژه/زیرپروژه است
        # ۳. آماده‌سازی داده‌های تنخواه (شماره، مبلغ، وضعیت، لینک جزئیات، لینک AJAX برای فاکتورها)
        # ۴. رندر تمپلیت جزئی `_ajax_level_tankhahs.html`
        # ۵. JsonResponse
        # برای این مثال، یک پاسخ ساده برمی‌گردانیم
        pba = get_object_or_404(ProjectBudgetAllocation.objects.select_related('project', 'subproject'), pk=pba_pk)
        target_name = pba.project.name if pba.project else ""
        if pba.subproject: target_name += f" / {pba.subproject.name}"

        tankhahs_data = []  # شما باید این را از دیتابیس پر کنید
        # مثال:
        # for t in Tankhah.objects.filter(project_budget_allocation=pba):
        #     tankhahs_data.append({'id': t.pk, 'number': t.number, ...})

        html_content = f"<p class='p-3 text-info small'>لیست تنخواه‌های مربوط به تخصیص پروژه <strong>{target_name}</strong> (PBA ID: {pba_pk}) در اینجا با AJAX بارگذاری می‌شود.</p>"
        if not tankhahs_data:
            html_content += "<p class='text-muted small ps-3'><em>هیچ تنخواهی برای این تخصیص پروژه یافت نشد.</em></p>"
        html_content += f"<a href='#' class='btn btn-sm btn-success ms-3 no-print'><i class='fas fa-plus-circle me-1'></i> افزودن تنخواه به این تخصیص</a>"
        return JsonResponse({'html_content': html_content, 'status': 'success'})
class APITankhahsForProjectView(LoginRequiredMixin, View):  # اگر تنخواه مستقیم به پروژه لینک است
    def get(self, request, project_pk, *args, **kwargs):
        # مشابه APITankhahsForPBAView اما فیلتر بر اساس project=project_pk
        project = get_object_or_404(Project, pk=project_pk)
        html_content = f"<p class='p-3 text-info small'>لیست تنخواه‌های مربوط به پروژه <strong>{project.name}</strong> در اینجا با AJAX بارگذاری می‌شود.</p>"
        html_content += f"<a href='#' class='btn btn-sm btn-success ms-3 no-print'><i class='fas fa-plus-circle me-1'></i> افزودن تنخواه به این پروژه</a>"
        return JsonResponse({'html_content': html_content, 'status': 'success'})
# --- ویوهای API برای سطوح پایین‌تر ( Factor) باید به همین ترتیب ایجاد شوند ---
class old_APIFactorsForTankhahView(LoginRequiredMixin, View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        # ۱. tankhah_instance = get_object_or_404(Tankhah, pk=tankhah_pk)
        # ۲. factors = Factor.objects.filter(tankhah=tankhah_instance)
        # ۳. آماده‌سازی داده‌های فاکتور (شماره، مبلغ، وضعیت، لینک جزئیات)
        # ۴. رندر تمپلیت جزئی `_ajax_level_factors.html`
        # ۵. JsonResponse
        from tankhah.models import Tankhah
        tankhah = get_object_or_404(Tankhah, pk=tankhah_pk)
        html_content = f"<p class='p-3 text-info small'>لیست فاکتورهای مربوط به تنخواه <strong>{tankhah.number}</strong> در اینجا با AJAX بارگذاری می‌شود.</p>"
        html_content += f"<a href='#' class='btn btn-sm btn-success ms-3 no-print'><i class='fas fa-plus-circle me-1'></i> افزودن فاکتور به این تنخواه</a>"
        return JsonResponse({'html_content': html_content, 'status': 'success'})
"""
شرح تک خطی: این API، لیست سازمان‌ها و خلاصه‌ی تخصیص‌ها برای یک دوره بودجه مشخص رو برمی‌گردونه.
تمپلیت مرتبط: reports/partials/_ajax_level_organizations.html (این HTML رو در پاسخ JSON برمی‌گردونه).
نحوه فراخوانی API‌ها در این تمپلیت:
سطرهای سازمان: هر سطر <tr> در این تمپلیت (مثل data-ajax-load-url="{{ org.budget_items_ajax_url }}")، API BudgetItemsForOrgPeriodAPIView رو برای بارگذاری سرفصل‌های بودجه مرتبط با اون سازمان و دوره فراخوانی می‌کنه.
"""
class  APIOrganizationsForPeriodView(LoginRequiredMixin, View):
    """
    API View برای دریافت خلاصه‌ای از سازمان‌ها در یک دوره بودجه خاص.
    این ویو برای بارگذاری محتوای سطح دوم (سازمان‌ها) در گزارش جامع استفاده می‌شود.
    """
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API request received: Get organizations for Period PK={period_pk}")
        try:
            period_instance = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            logger.debug(f"Found active BudgetPeriod: {period_instance.name}")

            # کوئری بهینه شده برای خلاصه سازمان‌ها
            org_summaries_qs = BudgetAllocation.objects.filter(
                budget_period=period_instance, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org_from_this_period=Sum('allocated_amount'),
                num_budget_items_for_org=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            organizations_data = []
            for org_sum_data in org_summaries_qs:
                org_id = org_sum_data['organization__id']

                # محاسبه مصرف سازمان (می‌توان با Subquery در بالا بهینه کرد)
                consumed_by_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_id,
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned_to_this_org = BudgetTransaction.objects.filter(
                    allocation__budget_period=period_instance,
                    allocation__organization_id=org_id,
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed_by_this_org_val = consumed_by_this_org - returned_to_this_org

                total_allocated_val = org_sum_data['total_allocated_to_org_from_this_period']
                remaining_for_this_org_val = total_allocated_val - net_consumed_by_this_org_val

                organizations_data.append({
                    'id': org_id,
                    'name': org_sum_data['organization__name'],
                    'code': org_sum_data['organization__code'],
                    'total_allocated': total_allocated_val,
                    'total_allocated_formatted': f"{total_allocated_val:,.0f}",
                    'net_consumed': net_consumed_by_this_org_val,
                    'net_consumed_formatted': f"{net_consumed_by_this_org_val:,.0f}",
                    'remaining': remaining_for_this_org_val,
                    'remaining_formatted': f"{remaining_for_this_org_val:,.0f}",
                    'num_budget_items': org_sum_data['num_budget_items_for_org'],
                    'utilization_percentage': (net_consumed_by_this_org_val / total_allocated_val * 100) \
                        if total_allocated_val > 0 else Decimal('0'),
                    # URL برای بارگذاری سرفصل‌های بودجه (BA) برای این سازمان و دوره
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',  # نام ویو شما در بالا
                                                      kwargs={'period_pk': period_pk, 'org_pk': org_id})
                })

            if not organizations_data:
                logger.info(f"No active organizations with allocations found for Period '{period_instance.name}'.")
                html_content = render_to_string('reports/partials/_no_organizations_found.html',
                                                {'period': period_instance})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            logger.info(f"Successfully prepared {len(organizations_data)} organization summaries for Period '{period_instance.name}'.")

            # رندر کردن تمپلیت جزئی برای محتوای HTML
            html_content = render_to_string(
                'reports/partials/_ajax_level_organizations.html',
                {
                    'organizations': organizations_data,
                    'parent_period_pk': period_pk,
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})

        except Http404 as e:
            logger.warning(f"API request failed: {e} (Period PK={period_pk})")
            return JsonResponse({'html_content': f'<p class="text-danger text-center small py-3"><em>{e}</em></p>',
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (Organizations for Period) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse(
                {'error': _("خطا در پردازش درخواست. لطفاً با پشتیبانی تماس بگیرید."), 'status': 'error'}, status=500)
