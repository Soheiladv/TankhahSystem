# reports/views.py
import logging
from decimal import Decimal
import jdatetime

from django.db.models import Sum, Q, Count, OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, View
from django.contrib import messages

# مدل‌های برنامه
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from core.PermissionBase import PermissionBaseView
from core.models import Organization, Project
from tankhah.models import Tankhah, Factor

# توابع محاسباتی
from budgets.budget_calculations import get_tankhah_used_budget

from django.db.models import Sum, Count, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views.generic import ListView
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils import formats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import jdatetime
import logging
from decimal import Decimal
from budgets.models import BudgetPeriod, BudgetAllocation, BudgetTransaction
from django.db.models.expressions import OuterRef, Subquery

# کتابخانه‌های PDF و Excel
# try:
#     from weasyprint import HTML, CSS
#
#     WEASYPRINT_AVAILABLE = True
# except ImportError:
#     WEASYPRINT_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# تعریف لاگر در سطح ماژول
logger = logging.getLogger(__name__)
class ComprehensiveBudgetReportView___(PermissionBaseView, ListView):
    """
    نمای اصلی گزارش جامع که لیست دوره‌های بودجه فعال را به همراه خلاصه‌ای از وضعیت آن‌ها نمایش می‌دهد.
    این ویو با لاگ‌های دقیق برای عیب‌یابی عدم نمایش داده‌ها مجهز شده است.
    """
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5  # فعال کردن صفحه‌بندی

    def get_queryset(self):
        logger.info("=" * 25 + " ComprehensiveBudgetReportView: Starting get_queryset " + "=" * 25)
        try:
            # Subquery ها برای محاسبات بهینه
            total_allocated_subquery = BudgetAllocation.objects.filter(budget_period_id=OuterRef('pk'),
                                                                       is_active=True).values(
                'budget_period_id').annotate(total=Sum('allocated_amount')).values('total')
            total_consumed_subquery = BudgetTransaction.objects.filter(allocation__budget_period_id=OuterRef('pk'),
                                                                       transaction_type='CONSUMPTION').values(
                'allocation__budget_period_id').annotate(total=Sum('amount')).values('total')
            total_returned_subquery = BudgetTransaction.objects.filter(allocation__budget_period_id=OuterRef('pk'),
                                                                       transaction_type='RETURN').values(
                'allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            base_filters = Q(is_active=True) & Q(is_completed=False)
            logger.debug(f"فیلترهای پایه برای BudgetPeriod: {base_filters}")

            queryset = BudgetPeriod.objects.filter(base_filters).select_related('organization').annotate(
                total_allocated_in_period=Coalesce(Subquery(total_allocated_subquery), Decimal('0'),
                                                   output_field=DecimalField()),
                total_consumed_in_period=Coalesce(Subquery(total_consumed_subquery), Decimal('0'),
                                                  output_field=DecimalField()),
                total_returned_in_period=Coalesce(Subquery(total_returned_subquery), Decimal('0'),
                                                  output_field=DecimalField()),
            ).order_by('-start_date', 'name')

            logger.info(f"تعداد دوره‌های بودجه یافت شده قبل از فیلتر جستجو: {queryset.count()}")

            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"اعمال فیلتر جستجو بر اساس نام دوره: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)

            final_count = queryset.count()
            logger.info(f"تعداد نهایی دوره‌های بودجه برای نمایش: {final_count}")
            if final_count == 0:
                logger.warning(
                    "مهم: هیچ دوره بودجه‌ای با فیلترهای مشخص شده یافت نشد. لطفاً از وجود دوره‌هایی با is_active=True و is_completed=False در دیتابیس اطمینان حاصل کنید.")

            return queryset
        except Exception as e:
            logger.error(f"خطای بحرانی در get_queryset رخ داد: {e}", exc_info=True)
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        logger.info("=" * 25 + " ComprehensiveBudgetReportView: Starting get_context_data " + "=" * 25)
        try:
            context = super().get_context_data(**kwargs)
            initial_periods = context.get('object_list', [])
            logger.debug(f"تعداد دوره‌ها برای پردازش در context: {len(initial_periods)}")

            processed_data = []
            for period in initial_periods:
                logger.debug(f"پردازش داده‌ها برای دوره: '{period.name}' (PK: {period.pk})")
                net_consumed = period.total_consumed_in_period - period.total_returned_in_period
                remaining_vs_total = period.total_amount - net_consumed
                utilization_vs_total = (net_consumed / period.total_amount * 100) if period.total_amount > 0 else 0

                processed_data.append({
                    'period': period,
                    'summary': {
                        'total_allocated_from_period': period.total_allocated_in_period,
                        'net_consumed_from_period': net_consumed,
                        'remaining_vs_total_amount': remaining_vs_total,
                        'utilization_percentage': utilization_vs_total,
                        'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime('%Y/%m/%d'),
                        'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime('%Y/%m/%d'),
                        'organizations_ajax_url': reverse('api_organizations_for_period',
                                                          kwargs={'period_pk': period.pk}),
                    }
                })

            context[self.context_object_name] = processed_data
            context['report_main_title'] = _("گزارش جامع بودجه")
            context['current_search_period'] = self.request.GET.get('search_period', '')

            logger.info(f"پردازش get_context_data با موفقیت انجام شد. تعداد نهایی آیتم‌ها: {len(processed_data)}")
            if not processed_data and not self.request.GET.get('search_period'):
                messages.info(self.request, _("هیچ دوره بودجه فعالی برای نمایش یافت نشد."))

            return context
        except Exception as e:
            logger.error(f"خطای بحرانی در get_context_data: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در پردازش داده‌های صفحه رخ داده است. لطفاً لاگ سرور را بررسی کنید."))
            context = super().get_context_data(**kwargs)
            context[self.context_object_name] = []
            return context
class APIBudgetItemsForOrgPeriodView__(PermissionBaseView, View):
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت سرفصل‌ها برای سازمان {org_pk} در دوره {period_pk}")
        try:
            consumed_subquery = BudgetTransaction.objects.filter(allocation_id=OuterRef('pk'),
                                                                 transaction_type='CONSUMPTION').values(
                'allocation_id').annotate(total=Sum('amount')).values('total')
            returned_subquery = BudgetTransaction.objects.filter(allocation_id=OuterRef('pk'),
                                                                 transaction_type='RETURN').values(
                'allocation_id').annotate(total=Sum('amount')).values('total')

            allocations_qs = BudgetAllocation.objects.filter(
                budget_period_id=period_pk, organization_id=org_pk, is_active=True
            ).select_related(
                'budget_item', 'project', 'subproject'
            ).annotate(
                consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('budget_item__name', 'project__name', 'subproject__name')

            logger.debug(f"تعداد تخصیص‌های یافت شده برای سازمان {org_pk}: {allocations_qs.count()}")
            if not allocations_qs.exists():
                return JsonResponse(
                    {'html_content': '<p class="text-center p-3">هیچ تخصیص فعالی برای این سازمان یافت نشد.</p>',
                     'status': 'empty'})

            response_data = []
            for ba in allocations_qs:
                net_consumed = ba.consumed - ba.returned
                remaining = ba.allocated_amount - net_consumed
                is_project_allocation = ba.project is not None
                target_name = ""
                if is_project_allocation:
                    target_name = ba.project.name
                    if ba.subproject:
                        target_name += f" / {ba.subproject.name}"

                response_data.append({
                    'is_project_allocation': is_project_allocation,
                    'pba_pk': ba.pk,
                    'budget_item_name': ba.budget_item.name if ba.budget_item else "نامشخص",
                    'budget_item_code': ba.budget_item.code if ba.budget_item else "-",
                    'project_display_name': target_name,
                    'allocated_formatted': f"{ba.allocated_amount:,.0f}",
                    'consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{remaining:,.0f}",
                    'utilization_percentage': (
                            net_consumed / ba.allocated_amount * 100) if ba.allocated_amount > 0 else 0,
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_pba',
                                                 kwargs={'pba_pk': ba.pk}) if is_project_allocation else None
                })

            html_content = render_to_string('reports/partials/_ajax_level_budget_allocations.html',
                                            {'allocations_data': response_data, 'parent_org_pk': org_pk})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Exception as e:
            logger.error(f"خطای بحرانی در APIBudgetItemsForOrgPeriodView (org_pk={org_pk}): {e}", exc_info=True)
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("خطا در پردازش درخواست.")}</p>',
                 'status': 'error'}, status=500)
class APITankhahsForPBAView__(PermissionBaseView, View):
    def get(self, request, pba_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت تنخواه‌ها برای تخصیص پروژه (PBA) PK={pba_pk}")
        try:
            pba_instance = get_object_or_404(BudgetAllocation.objects.select_related('project', 'subproject'),
                                             pk=pba_pk)

            tankhahs_qs = Tankhah.objects.filter(project_budget_allocation=pba_instance,
                                                 is_archived=False).select_related('created_by',
                                                                                   'current_stage').order_by('-date',
                                                                                                             '-pk')

            logger.debug(f"تعداد تنخواه‌های یافت شده برای PBA {pba_pk}: {tankhahs_qs.count()}")
            if not tankhahs_qs.exists():
                return JsonResponse(
                    {'html_content': '<p class="text-center p-3">هیچ تنخواهی برای این تخصیص یافت نشد.</p>',
                     'status': 'empty'})

            tankhahs_data = []
            for tankhah in tankhahs_qs:
                used_amount = get_tankhah_used_budget(tankhah)
                tankhahs_data.append({
                    'id': tankhah.pk,
                    'number': tankhah.number,
                    'amount_formatted': f"{tankhah.amount:,.0f}",
                    'used_formatted': f"{used_amount:,.0f}",
                    'status_display': tankhah.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=tankhah.date).strftime('%Y/%m/%d'),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': tankhah.pk}),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': tankhah.pk}),
                })

            parent_pba_name = pba_instance.project.name if pba_instance.project else ""
            if pba_instance.subproject:
                parent_pba_name += f" / {pba_instance.subproject.name}"

            html_content = render_to_string('reports/partials/_ajax_level_tankhahs.html',
                                            {'tankhahs_list_data': tankhahs_data, 'parent_pba_pk': pba_pk,
                                             'parent_pba_name': parent_pba_name})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            logger.warning(f"API Error: تخصیص بودجه پروژه با PK={pba_pk} یافت نشد.")
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("تخصیص بودجه یافت نشد.")}</p>',
                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"خطای بحرانی در APITankhahsForPBAView (pba_pk={pba_pk}): {e}", exc_info=True)
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("خطا در پردازش درخواست.")}</p>',
                 'status': 'error'}, status=500)
class APIFactorsForTankhahView__(PermissionBaseView, View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت فاکتورها برای تنخواه PK={tankhah_pk}")
        try:
            tankhah_instance = get_object_or_404(Tankhah.objects.select_related('organization', 'project'),
                                                 pk=tankhah_pk)

            factors_qs = Factor.objects.filter(tankhah=tankhah_instance).select_related('category',
                                                                                        'created_by').order_by('-date')

            logger.debug(f"تعداد فاکتورهای یافت شده برای تنخواه {tankhah_pk}: {factors_qs.count()}")
            if not factors_qs.exists():
                return JsonResponse(
                    {'html_content': '<p class="text-center p-3">هیچ فاکتوری برای این تنخواه یافت نشد.</p>',
                     'status': 'empty'})

            factors_data = []
            for factor in factors_qs:
                factors_data.append({
                    'id': factor.pk,
                    'number': factor.number,
                    'amount_formatted': f"{factor.amount:,.0f}",
                    'status_display': factor.get_status_display(),
                    'category_name': factor.category.name if factor.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=factor.date).strftime('%Y/%m/%d'),
                    'detail_url': reverse('factor_detail', kwargs={'factor_pk': factor.pk}),
                })

            html_content = render_to_string('reports/partials/_ajax_level_factors.html',
                                            {'factors_list_data': factors_data, 'parent_tankhah_pk': tankhah_pk,
                                             'parent_tankhah_number': tankhah_instance.number})
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            logger.warning(f"API Error: تنخواه با PK={tankhah_pk} یافت نشد.")
            return JsonResponse({'html_content': f'<p class="text-danger p-3 text-center">{_("تنخواه یافت نشد.")}</p>',
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"خطای بحرانی در APIFactorsForTankhahView (tankhah_pk={tankhah_pk}): {e}", exc_info=True)
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("خطا در پردازش درخواست.")}</p>',
                 'status': 'error'}, status=500)
class APIBudgetItemsForOrgPeriodView__APIBudgetItemsForOrgPeriodView(PermissionBaseView, View):
    """
    این ویو تخصیص‌های اصلی (سرفصل) و تخصیص‌های فرزند (پروژه) را نمایش می‌دهد.
    """

    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت سرفصل‌ها و پروژه‌ها برای سازمان {org_pk} در دوره {period_pk}")

        # ۱. واکشی تخصیص‌های اصلی (main_allocations)
        main_allocations = BudgetAllocation.objects.filter(
            budget_period_id=period_pk,
            organization_id=org_pk,
            project__isnull=True,  # مهم: فقط تخصیص‌های اصلی
            is_active=True
        ).select_related('budget_item').order_by('budget_item__name')

        if not main_allocations.exists():
            return JsonResponse(
                {'html_content': '<p class="text-center p-3">هیچ سرفصل بودجه‌ای یافت نشد.</p>', 'status': 'empty'})

        # بهینه سازی: تمام پروژه‌های فرزند را در یک کوئری واکشی می‌کنیم
        main_alloc_ids = main_allocations.values_list('pk', flat=True)
        project_allocations = BudgetAllocation.objects.filter(
            budget_allocation_id__in=main_alloc_ids,
            is_active=True
        ).select_related('project', 'subproject', 'budget_allocation')

        # دسته‌بندی پروژه‌ها بر اساس والد
        projects_by_parent = {}
        for pba in project_allocations:
            if pba.budget_allocation_id not in projects_by_parent:
                projects_by_parent[pba.budget_allocation_id] = []
            projects_by_parent[pba.budget_allocation_id].append(pba)

        response_data = []
        for main_alloc in main_allocations:
            child_projects = projects_by_parent.get(main_alloc.pk, [])

            project_data = []
            for pba in child_projects:
                # محاسبات مصرف را می‌توان به API بعدی موکول کرد یا با annotate بهینه انجام داد
                project_data.append({
                    'pba_pk': pba.pk,
                    'project_name': f"{pba.project.name}{f' / {pba.subproject.name}' if pba.subproject else ''}",
                    'allocated_formatted': f"{pba.allocated_amount:,.0f}",
                    # مهم: این URL به API بعدی با PK صحیح (فرزند) اشاره دارد
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_pba', kwargs={'pba_pk': pba.pk})
                })

            response_data.append({
                'ba_pk': main_alloc.pk,
                'budget_item_display': f"کد: {main_alloc.budget_item.code or '-'} - {main_alloc.budget_item.name}",
                'allocated_formatted': f"{main_alloc.allocated_amount:,.0f}",
                'project_allocations': project_data,
            })

        html_content = render_to_string(
            'reports/partials/_ajax_level_budget_allocations.html',
            {'allocations_data': response_data}
        )
        return JsonResponse({'html_content': html_content, 'status': 'success'})#---------------------------------------------------------------------
class __APIFactorsForTankhahView(PermissionBaseView, View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API request: Get Factors for Tankhah PK={tankhah_pk}")
        try:
            tankhah = get_object_or_404(Tankhah.objects.select_related('organization', 'project'), pk=tankhah_pk)
            factors = Factor.objects.filter(tankhah=tankhah).select_related('category', 'created_by').prefetch_related('items').order_by('-date')

            if not factors.exists():
                html_content = render_to_string('reports/partials/_no_factors_found.html', {'parent_tankhah': tankhah})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            factors_data = []
            for f in factors:
                items = f.items.all()
                item_data = [
                    {
                        'description': item.description,
                        'amount_formatted': f"{item.amount or 0:,.0f}",
                        'quantity': item.quantity,
                        'unit_price': f"{item.unit_price or 0:,.0f}",
                        'status_display': item.get_status_display(),
                    } for item in items
                ]

                factors_data.append({
                    'id': f.pk,
                    'number': f.number,
                    'amount_formatted': f"{f.amount or 0:,.0f}",
                    'status_display': f.get_status_display(),
                    'category_name': f.category.name if f.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime('%Y/%m/%d') if f.date else "-",
                    'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                    'items': item_data,
                    'items_count': len(item_data),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_factors.html',
                {'factors_list_data': factors_data, 'parent_tankhah_pk': tankhah_pk, 'parent_tankhah_number': tankhah.number}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تنخواه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش فاکتورها.")}),
                 'status': 'error'},
                status=500
            )

class __APIOrganizationsForPeriodView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API request: Get organizations for Period PK={period_pk}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)

            org_summaries = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='RETURN'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('organization__name')

            organizations_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'code': org['organization__code'],
                    'total_allocated_formatted': f"{org['total_allocated']:,.0f}",
                    'net_consumed_formatted': f"{org['consumed'] - org['returned']:,.0f}",
                    'remaining_formatted': f"{org['total_allocated'] - (org['consumed'] - org['returned']):,.0f}",
                    'num_budget_items': org['num_budget_items'],
                    'utilization_percentage': (
                            (org['consumed'] - org['returned']) / org['total_allocated'] * 100
                    ) if org['total_allocated'] > 0 else 0,
                    'budget_items_ajax_url': reverse(
                        'api_budget_items_for_org_period',
                        kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']}
                    )
                } for org in org_summaries
            ]

            if not organizations_data:
                html_content = render_to_string('reports/partials/_no_organizations_found.html', {'period': period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            html_content = render_to_string(
                'reports/partials/_ajax_level_organizations.html',
                {'organizations': organizations_data, 'parent_period_pk': period_pk}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404 as e:
            logger.warning(f"Period PK={period_pk} not found: {e}")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_organizations_found.html',
                                                  {'message': _("دوره بودجه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                  {'error_message': _("خطا در پردازش سازمان‌ها.")}),
                 'status': 'error'},
                status=500
            )

#---------------------------------------------------------------------
class __ComprehensiveBudgetReportView(PermissionBaseView, ListView):
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Starting get_queryset...")
        try:
            total_allocated_subquery = BudgetAllocation.objects.filter(
                budget_period_id=OuterRef('pk'), is_active=True
            ).values('budget_period_id').annotate(total=Sum('allocated_amount')).values('total')

            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='CONSUMPTION'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            returned_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='RETURN'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            queryset = BudgetPeriod.objects.filter(
                is_active=True, is_completed=False
            ).select_related('organization').annotate(
                period_total_allocated=Coalesce(Subquery(total_allocated_subquery), Decimal('0'),
                                                output_field=DecimalField()),
                period_total_consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('-start_date', 'name')

            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"Applying search filter: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)

            count = queryset.count()
            logger.info(f"Found {count} BudgetPeriods.")
            if count == 0:
                logger.warning("No active BudgetPeriods found.")
            return queryset
        except Exception as e:
            logger.error(f"Error in get_queryset: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در بارگذاری دوره‌های بودجه رخ داد."))
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        logger.info(f"[{self.__class__.__name__}] - Starting get_context_data...")
        context = super().get_context_data(**kwargs)
        periods = context.get('object_list', [])

        processed_data = []
        for period in periods:
            net_consumed = period.period_total_consumed - period.period_total_returned
            remaining = period.total_amount - net_consumed
            utilization = (net_consumed / period.total_amount * 100) if period.total_amount else 0

            org_summaries = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                org_total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                org_consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                org_returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='RETURN'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('organization__name')

            org_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'code': org['organization__code'],
                    'total_allocated': org['org_total_allocated'] or 0,
                    'total_allocated_formatted': f"{org['org_total_allocated'] or 0:,.0f}",
                    'net_consumed': (org['org_consumed'] - org['org_returned']) or 0,
                    'net_consumed_formatted': f"{(org['org_consumed'] - org['org_returned']) or 0:,.0f}",
                    'remaining': (org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0,
                    'remaining_formatted': f"{(org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0:,.0f}",
                    'num_budget_items': org['num_budget_items'],
                    'utilization_percentage': (
                            (org['org_consumed'] - org['org_returned']) / org['org_total_allocated'] * 100
                    ) if org['org_total_allocated'] else 0,
                    'budget_items_ajax_url': reverse(
                        'api_budget_items_for_org_period',
                        kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']}
                    )
                } for org in org_summaries
            ]

            processed_data.append({
                'period': period,
                'summary': {
                    'total_budget': period.total_amount or 0,
                    'total_allocated': period.period_total_allocated or 0,
                    'net_consumed': net_consumed or 0,
                    'remaining': remaining or 0,
                    'utilization_percentage': utilization,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime(
                        '%Y/%m/%d') if period.start_date else '-',
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime(
                        '%Y/%m/%d') if period.end_date else '-',
                },
                'organization_summaries': org_data
            })

        context[self.context_object_name] = processed_data
        context['report_main_title'] = _("گزارش جامع بودجه")
        context['current_search_period'] = self.request.GET.get('search_period', '')

        if not processed_data:
            messages.info(self.request, _("هیچ دوره بودجه فعالی یافت نشد."))
        return context

    def generate_excel_report(self, report_data):
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl is not installed.")
            messages.error(self.request, _("لطفاً openpyxl را نصب کنید."))
            return None

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="budget_report.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True

        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل (ریال)"),
            _("تخصیص (ریال)"), _("مصرف (ریال)"), _("مانده (ریال)"), _("% مصرف"),
            _("سازمان تخصیص"), _("کد"), _("تخصیص سازمان (ریال)"), _("مصرف سازمان (ریال)"),
            _("مانده سازمان (ریال)"), _("% مصرف سازمان")
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col)].width = 20

        row = 2
        for period_data in report_data:
            period = period_data['period']
            summary = period_data['summary']
            start_row = row

            if not period_data['organization_summaries']:
                sheet.cell(row=row, column=1, value=period.name)
                sheet.cell(row=row, column=2, value=period.organization.name)
                sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                row += 1
            else:
                for org in period_data['organization_summaries']:
                    sheet.cell(row=row, column=1, value=period.name)
                    sheet.cell(row=row, column=2, value=period.organization.name)
                    sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                    sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                    sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                    sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                    sheet.cell(row=row, column=10, value=org['name'])
                    sheet.cell(row=row, column=11, value=org['code'])
                    sheet.cell(row=row, column=12, value=org['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=13, value=org['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=14, value=org['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=15, value=org['utilization_percentage'] / 100).number_format = '0.0%'
                    row += 1

                if len(period_data['organization_summaries']) > 1:
                    for col in range(1, 10):
                        sheet.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)

        workbook.save(response)
        return response

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data = context.get(self.context_object_name, [])

        if output_format == 'excel':
            response = self.generate_excel_report(report_data)
            if response:
                return response
            messages.error(self.request, _("خطا در تولید فایل اکسل."))
            return super().render_to_response(context, **response_kwargs)

        return super().render_to_response(context, **response_kwargs)
#---------------------------------------------------------------------
class ComprehensiveBudgetReportView(PermissionBaseView, ListView):
    model = BudgetPeriod
    template_name = 'reports/v2/comprehensive_report_main.html'
    context_object_name = 'budget_periods_data'
    paginate_by = 5

    def get_queryset(self):
        logger.info(f"[{self.__class__.__name__}] - Starting get_queryset...")
        try:
            total_allocated_subquery = BudgetAllocation.objects.filter(
                budget_period_id=OuterRef('pk'), is_active=True
            ).values('budget_period_id').annotate(total=Sum('allocated_amount')).values('total')

            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='CONSUMPTION'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            returned_subquery = BudgetTransaction.objects.filter(
                allocation__budget_period_id=OuterRef('pk'), transaction_type='RETURN'
            ).values('allocation__budget_period_id').annotate(total=Sum('amount')).values('total')

            queryset = BudgetPeriod.objects.filter(
                is_active=True, is_completed=False
            ).select_related('organization').annotate(
                period_total_allocated=Coalesce(Subquery(total_allocated_subquery), Decimal('0'),
                                                output_field=DecimalField()),
                period_total_consumed=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                period_total_returned=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('-start_date', 'name')

            search_query = self.request.GET.get('search_period')
            if search_query:
                logger.debug(f"Applying search filter: '{search_query}'")
                queryset = queryset.filter(name__icontains=search_query)

            count = queryset.count()
            logger.info(f"Found {count} BudgetPeriods.")
            if count == 0:
                logger.warning("No active BudgetPeriods found.")
            return queryset
        except Exception as e:
            logger.error(f"Error in get_queryset: {e}", exc_info=True)
            messages.error(self.request, _("خطایی در بارگذاری دوره‌های بودجه رخ داد."))
            return BudgetPeriod.objects.none()

    def get_context_data(self, **kwargs):
        logger.info(f"[{self.__class__.__name__}] - Starting get_context_data...")
        context = super().get_context_data(**kwargs)
        periods = context.get('object_list', [])

        processed_data = []
        for period in periods:
            net_consumed = period.period_total_consumed - period.period_total_returned
            remaining = period.total_amount - net_consumed
            utilization = (net_consumed / period.total_amount * 100) if period.total_amount else 0

            org_summaries = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                org_total_allocated=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True),
                org_consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                org_returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation__budget_period=period,
                            allocation__organization_id=OuterRef('organization__id'),
                            transaction_type='RETURN'
                        ).values('allocation__organization_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('organization__name')

            org_data = [
                {
                    'id': org['organization__id'],
                    'name': org['organization__name'],
                    'code': org['organization__code'],
                    'total_allocated': org['org_total_allocated'] or 0,
                    'total_allocated_formatted': f"{org['org_total_allocated'] or 0:,.0f}",
                    'net_consumed': (org['org_consumed'] - org['org_returned']) or 0,
                    'net_consumed_formatted': f"{(org['org_consumed'] - org['org_returned']) or 0:,.0f}",
                    'remaining': (org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0,
                    'remaining_formatted': f"{(org['org_total_allocated'] - (org['org_consumed'] - org['org_returned'])) or 0:,.0f}",
                    'num_budget_items': org['num_budget_items'],
                    'utilization_percentage': (
                        (org['org_consumed'] - org['org_returned']) / org['org_total_allocated'] * 100
                    ) if org['org_total_allocated'] else 0,
                    'budget_items_ajax_url': reverse(
                        'api_budget_items_for_org_period',
                        kwargs={'period_pk': period.pk, 'org_pk': org['organization__id']}
                    )
                } for org in org_summaries
            ]

            processed_data.append({
                'period': period,
                'summary': {
                    'total_budget': period.total_amount or 0,
                    'total_allocated': period.period_total_allocated or 0,
                    'net_consumed': net_consumed or 0,
                    'remaining': remaining or 0,
                    'utilization_percentage': utilization,
                    'start_date_jalali': jdatetime.date.fromgregorian(date=period.start_date).strftime(
                        '%Y/%m/%d') if period.start_date else '-',
                    'end_date_jalali': jdatetime.date.fromgregorian(date=period.end_date).strftime(
                        '%Y/%m/%d') if period.end_date else '-',
                },
                'organization_summaries': org_data
            })

        context[self.context_object_name] = processed_data
        context['report_main_title'] = _("گزارش جامع بودجه")
        context['current_search_period'] = self.request.GET.get('search_period', '')

        if not processed_data:
            messages.info(self.request, _("هیچ دوره بودجه فعالی یافت نشد."))
        return context

    def generate_excel_report(self, report_data):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl is not installed.")
            messages.error(self.request, _("لطفاً openpyxl را نصب کنید."))
            return None

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="budget_report.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _("گزارش جامع بودجه")
        sheet.sheet_view.rightToLeft = True

        headers = [
            _("دوره بودجه"), _("سازمان دوره"), _("تاریخ شروع"), _("تاریخ پایان"), _("بودجه کل (ریال)"),
            _("تخصیص (ریال)"), _("مصرف (ریال)"), _("مانده (ریال)"), _("% مصرف"),
            _("سازمان تخصیص"), _("کد"), _("تخصیص سازمان (ریال)"), _("مصرف سازمان (ریال)"),
            _("مانده سازمان (ریال)"), _("% مصرف سازمان")
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col)].width = 20

        row = 2
        for period_data in report_data:
            period = period_data['period']
            summary = period_data['summary']
            start_row = row

            if not period_data['organization_summaries']:
                sheet.cell(row=row, column=1, value=period.name)
                sheet.cell(row=row, column=2, value=period.organization.name)
                sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                row += 1
            else:
                for org in period_data['organization_summaries']:
                    sheet.cell(row=row, column=1, value=period.name)
                    sheet.cell(row=row, column=2, value=period.organization.name)
                    sheet.cell(row=row, column=3, value=summary['start_date_jalali'])
                    sheet.cell(row=row, column=4, value=summary['end_date_jalali'])
                    sheet.cell(row=row, column=5, value=summary['total_budget']).number_format = '#,##0'
                    sheet.cell(row=row, column=6, value=summary['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=7, value=summary['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=8, value=summary['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=9, value=summary['utilization_percentage'] / 100).number_format = '0.0%'
                    sheet.cell(row=row, column=10, value=org['name'])
                    sheet.cell(row=row, column=11, value=org['code'])
                    sheet.cell(row=row, column=12, value=org['total_allocated']).number_format = '#,##0'
                    sheet.cell(row=row, column=13, value=org['net_consumed']).number_format = '#,##0'
                    sheet.cell(row=row, column=14, value=org['remaining']).number_format = '#,##0'
                    sheet.cell(row=row, column=15, value=org['utilization_percentage'] / 100).number_format = '0.0%'
                    row += 1

                if len(period_data['organization_summaries']) > 1:
                    for col in range(1, 10):
                        sheet.merge_cells(start_row=start_row, start_column=col, end_row=row - 1, end_column=col)

        workbook.save(response)
        return response

    def render_to_response(self, context, **response_kwargs):
        output_format = self.request.GET.get('output_format', 'html').lower()
        report_data = context.get(self.context_object_name, [])

        if output_format == 'excel':
            response = self.generate_excel_report(report_data)
            if response:
                return response
            messages.error(self.request, _("خطا در تولید فایل اکسل."))
            return super().render_to_response(context, **response_kwargs)

        return super().render_to_response(context, **response_kwargs)
#---------------------------------------------------------------------
class ____BudgetItemsForOrgPeriodAPIView(PermissionBaseView, View):
    """
    API View برای دریافت جزئیات BudgetAllocations (سرفصل‌ها) و
    ProjectBudgetAllocations مرتبط برای یک سازمان و دوره بودجه خاص.
    این ویو برای پاسخ به درخواست‌های AJAX از گزارش جامع استفاده می‌شود.
    """
    print(f'BudgetItemsForOrgPeriodAPIView is load ')

    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request received: Get budget items for Period PK={period_pk}, Org PK={org_pk}")
        try:
            # واکشی دوره بودجه و سازمان با بررسی فعال بودن
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.debug(f"Found active BudgetPeriod: {budget_period.name} and Organization: {organization.name}")

            # ۱. واکشی BudgetAllocation ها (سرفصل‌ها) برای این سازمان و دوره
            # budget_allocations_qs = BudgetAllocation.objects.filter(
            #     budget_period=budget_period,
            #     organization=organization,
            #     is_active=True  # فقط تخصیص‌های فعال
            # ).select_related('budget_item',  # برای نام و کد سرفصل
            #                  # 'project' # اگر BudgetAllocation مستقیماً به یک پروژه کلی هم می‌تواند لینک شود
            #                  ).prefetch_related(Prefetch(
            #     'project_allocations',  # related_name از BudgetAllocation به ProjectBudgetAllocation
            #     queryset= BudgetAllocation.objects.filter(is_active=True)
            #     .select_related('project', 'subproject')  # برای نام پروژه و زیرپروژه
            #     .order_by('project__name', 'subproject__name'),
            #     to_attr='active_project_allocations_list'  # نامی برای دسترسی در حلقه
            #  )
            # ).order_by('budget_item__name')  # مرتب‌سازی بر اساس نام سرفصل
            budget_allocations_qs = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related(
                'budget_item',
                'project',
                'subproject'
            ).order_by('budget_item__name', 'project__name', 'subproject__name')  # مرتب‌سازی کامل

            if not budget_allocations_qs.exists():
                logger.info(
                    f"No active BudgetAllocations found for Org '{organization.name}' in Period '{budget_period.name}'.")
                # می‌توانید یک رشته HTML یا JSON خالی با پیام مناسب برگردانید
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                                {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            # ۲. آماده‌سازی داده‌ها برای ارسال
            response_data_allocations = []
            for ba_instance in budget_allocations_qs:
                # محاسبه مصرف و مانده برای این BudgetAllocation خاص
                from reports.views import calculate_total_consumed_on_budget_allocation
                from reports.views import get_budget_allocation_remaining_amount
                consumed_on_this_ba = calculate_total_consumed_on_budget_allocation(ba_instance,
                                                                                    use_cache=False)  # False برای داده بروز در AJAX
                remaining_on_this_ba = get_budget_allocation_remaining_amount(ba_instance, use_cache=False)

                # آماده‌سازی لیست پروژه‌های تخصیص یافته از این BudgetAllocation (سرفصل)
                projects_under_this_ba = []
                # getattr برای دسترسی امن به to_attr از prefetch
                for pba in getattr(ba_instance, 'active_project_allocations_list', []):
                    project_name_display = pba.project.name if pba.project else _("پروژه نامشخص")
                    if pba.subproject:
                        project_name_display += f" / {pba.subproject.name}"

                    # مانده کلی پروژه (از تمام منابع) - این می‌تواند اختیاری باشد
                    project_overall_remaining_str = "-"
                    if pba.project:
                        try:
                            # این تابع ممکن است کوئری اضافی بزند، برای تعداد زیاد باید بهینه شود
                            from budgets.budget_calculations import get_project_remaining_budget
                            project_overall_remaining = get_project_remaining_budget(pba.project)
                            project_overall_remaining_str = f"{project_overall_remaining:,.0f}"
                        except Exception as e_proj_rem:
                            logger.warning(
                                f"Could not calculate overall remaining for project {pba.project.pk}: {e_proj_rem}")

                    projects_under_this_ba.append({
                        'pba_pk': pba.pk,
                        'project_name_display': project_name_display,
                        'allocated_to_pba': pba.allocated_amount,  # مبلغ تخصیص یافته به این پروژه/زیرپروژه از این سرفصل
                        'allocated_to_pba_formatted': f"{pba.allocated_amount:,.0f}",
                        'pba_detail_url': reverse('project_budget_allocation_detail', kwargs={'pk': pba.pk}),
                        'project_overall_remaining_formatted': project_overall_remaining_str,  # مانده کلی پروژه
                    })

                response_data_allocations.append({
                    'ba_pk': ba_instance.pk,
                    'budget_item_name': ba_instance.budget_item.name if ba_instance.budget_item else _("سرفصل نامشخص"),
                    'budget_item_code': ba_instance.budget_item.code if ba_instance.budget_item else "-",
                    'ba_allocated_amount': ba_instance.allocated_amount,
                    'ba_allocated_amount_formatted': f"{ba_instance.allocated_amount:,.0f}",
                    'ba_consumed_amount_formatted': f"{consumed_on_this_ba:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining_on_this_ba:,.0f}",
                    'ba_utilization_percentage': (
                            consumed_on_this_ba / ba_instance.allocated_amount * 100) if ba_instance.allocated_amount > 0 else 0,
                    'assigned_projects_details': projects_under_this_ba,  # لیست پروژه‌های زیرمجموعه این سرفصل
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': ba_instance.pk}),
                    # لینک به گزارش کامل این BudgetAllocation
                    # 'add_pba_url': reverse('project_budget_allocation') + f"?organization_id={ba_instance.pk}"
                    # لینک برای ایجاد PBA جدید: نیاز به organization_id دارد
                    'add_pba_url': reverse('project_budget_allocation', kwargs={
                        'organization_id': organization.pk}) + f"?budget_allocation_id={ba_instance.pk}"
                    # لینک برای ایجاد PBA جدید
                })

            logger.info(
                f"Successfully prepared {len(response_data_allocations)} BudgetAllocation details for Org '{organization.name}' in Period '{budget_period.name}'.")

            # رندر کردن تمپلیت جزئی برای محتوای HTML
            html_content = render_to_string(
                # 'reports/partials/_budget_items_for_org_ajax.html',  # تمپلیت partial جدید
                'reports/partials/_ajax_level_budget_allocations.html',  # تمپلیت partial جدید
                {
                    'budget_allocations_data': response_data_allocations,  # نام متغیر برای تمپلیت partial
                    'parent_period_pk': period_pk,  # برای id های منحصر به فرد آکاردئون
                    'parent_org_pk': org_pk
                }
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})

        except Http404 as e:  # اگر دوره یا سازمان یافت نشد
            logger.warning(f"API request failed: {e} (Period PK={period_pk}, Org PK={org_pk})")
            return JsonResponse({'html_content': f'<p class="text-danger text-center small py-3"><em>{e}</em></p>',
                                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error for budget items (Period PK={period_pk}, Org PK={org_pk}): {e}", exc_info=True)
            return JsonResponse(
                {'error': _("خطا در پردازش درخواست. لطفاً با پشتیبانی تماس بگیرید."), 'status': 'error'}, status=500)

class BudgetItemsForOrgPeriodAPIView(View):
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request received: Get budget items for Period PK={period_pk}, Org PK={org_pk}, User={request.user}, Path={request.path}")
        logger.debug(f"Request headers: {dict(request.headers)}")
        try:
            # دریافت دوره بودجه و سازمان
            logger.debug(f"Fetching BudgetPeriod (pk={period_pk}) and Organization (pk={org_pk})")
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.info(f"Found active BudgetPeriod: {budget_period.name} (ID={budget_period.pk}) and Organization: {organization.name} (ID={organization.pk})")

            # دریافت تخصیص‌های بودجه
            logger.debug(f"Querying BudgetAllocations for period={period_pk}, org={org_pk}")
            allocations = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').annotate(
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='RETURN'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('budget_item__name')

            logger.debug(f"Found {allocations.count()} BudgetAllocations: {[alloc.pk for alloc in allocations]}")

            # بررسی وجود تخصیص‌ها
            if not allocations.exists():
                logger.warning(f"No active BudgetAllocations found for Period '{budget_period.name}' and Organization '{organization.name}'")
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                               {'organization': organization, 'period': budget_period})
                logger.debug("Returning 'empty' response with no_budget_items_found template")
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            response_data = []
            for alloc in allocations:
                net_consumed = alloc.consumed - alloc.returned
                remaining = alloc.allocated_amount - net_consumed
                alloc_name = alloc.budget_item.name or 'سرفصل نامشخص'
                logger.debug(f"Processing BudgetAllocation ID={alloc.pk}, Name={alloc_name}, Consumed={net_consumed}, Remaining={remaining}")

                # تنخواه‌های متصل
                tankhahs = Tankhah.objects.filter(
                    budget_allocation=alloc,
                    is_archived=False
                ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')
                logger.debug(f"Found {tankhahs.count()} Tankhahs for BudgetAllocation ID={alloc.pk}: {[t.pk for t in tankhahs]}")

                tankhah_data = []
                for t in tankhahs:
                    factors = t.factors.all()
                    logger.debug(f"Processing Tankhah ID={t.pk}, Number={t.number}, Factors count={factors.count()}")
                    factor_data = [
                        {
                            'id': f.pk,
                            'number': f.number,
                            'amount_formatted': f"{f.amount or 0:,.0f}",
                            'status_display': f.get_status_display(),
                            'category_name': f.category.name if f.category else "-",
                            'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime(
                                '%Y/%m/%d') if f.date else "-",
                            'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                            'items_count': f.items.count(),
                            'items': [
                                {
                                    'description': item.description,
                                    'amount_formatted': f"{item.amount or 0:,.0f}",
                                    'quantity': item.quantity,
                                    'unit_price': f"{item.unit_price or 0:,.0f}",
                                    'status_display': item.get_status_display(),
                                } for item in f.items.all()
                            ]
                        } for f in factors
                    ]

                    tankhah_data.append({
                        'id': t.pk,
                        'number': t.number,
                        'amount_formatted': f"{t.amount or 0:,.0f}",
                        'status_display': t.get_status_display(),
                        'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime(
                            '%Y/%m/%d') if t.date else '-',
                        'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                        'factors': factor_data,
                        'factors_count': len(factor_data),
                        'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    })

                response_data.append({
                    'ba_pk': alloc.pk,
                    'budget_item_name': alloc_name,
                    'budget_item_code': alloc.budget_item.code or '-',
                    'ba_allocated_amount_formatted': f"{alloc.allocated_amount or 0:,.0f}",
                    'ba_consumed_amount_formatted': f"{net_consumed or 0:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining or 0:,.0f}",
                    'ba_utilization_percentage': (net_consumed / alloc.allocated_amount * 100) if alloc.allocated_amount else 0,
                    'tankhahs': tankhah_data,
                    'tankhahs_count': len(tankhah_data),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_allocation', kwargs={'alloc_pk': alloc.pk}),
                    'add_tankhah_url': reverse('tankhah_create') + f"?budget_allocation_id={alloc.pk}",
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': alloc.pk}),
                })

            logger.info(f"Successfully prepared {len(response_data)} BudgetAllocation details")
            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {'budget_allocations_data': response_data, 'parent_period_pk': period_pk, 'parent_org_pk': org_pk}
            )
            logger.debug(f"Rendered HTML content length: {len(html_content)} characters")
            return JsonResponse({'html_content': html_content, 'status': 'success'})

        except Http404 as e:
            logger.error(f"404 Error: {str(e)} for Period PK={period_pk}, Org PK={org_pk}")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': str(e)}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)} for Period PK={period_pk}, Org PK={org_pk}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                 {'error_message': _("خطا در پردازش سرفصل‌ها.")}),
                 'status': 'error'},
                status=500
            )
# ---------------------------------------------------------------------
class OrganizationAllocationsAPIView(PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization_summaries = []
            org_alloc_queryset = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org=Sum('allocated_amount'),
                num_budget_items=Count('budget_item', distinct=True)
            ).order_by('organization__name')

            for org_summary in org_alloc_queryset:
                # محاسبه مصرف (این بخش همچنان نیاز به بهینه‌سازی دارد اگر تعداد زیاد است)
                consumed = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='CONSUMPTION'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                returned = BudgetTransaction.objects.filter(
                    allocation__budget_period=period,
                    allocation__organization_id=org_summary['organization__id'],
                    transaction_type='RETURN'
                ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']
                net_consumed = consumed - returned

                organization_summaries.append({
                    'id': org_summary['organization__id'],
                    'name': org_summary['organization__name'],
                    'code': org_summary['organization__code'],
                    'total_allocated_formatted': f"{org_summary['total_allocated_to_org']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{org_summary['total_allocated_to_org'] - net_consumed:,.0f}",
                    'num_budget_items': org_summary['num_budget_items'],
                    'utilization_percentage': (net_consumed / org_summary['total_allocated_to_org'] * 100) if
                    org_summary['total_allocated_to_org'] > 0 else 0,
                    # URL برای بارگذاری سرفصل‌های این سازمان
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_summary['organization__id']})
                })
            return JsonResponse({'organizations': organization_summaries, 'status': 'success'})
        except Http404:
            return JsonResponse({'error': _("دوره بودجه یافت نشد."), 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"API Error (OrgAllocations) for Period PK={period_pk}: {e}", exc_info=True)
            return JsonResponse({'error': _("خطا در پردازش."), 'status': 'error'}, status=500)
#---------------------------------------------------------------------
#APIBudgetItemsForOrgPeriodView این ویو برای نمایش سلسله‌مراتب کامل (سرفصل‌ها → تخصیص‌های پروژه → تنخواه‌ها → فاکتورها) اصلاح می‌شود.
class APIBudgetItemsForOrgPeriodView(View):  # فرض می‌کنیم PermissionBaseView وجود دارد، در غیر این صورت حذف کنید
    def get(self, request, period_pk, org_pk, *args, **kwargs):
        logger.info(f"API request: Get budget items for Period PK={period_pk}, Org PK={org_pk}")
        try:
            # دریافت دوره بودجه و سازمان
            budget_period = get_object_or_404(BudgetPeriod, pk=period_pk, is_active=True, is_completed=False)
            organization = get_object_or_404(Organization, pk=org_pk, is_active=True)
            logger.info(f"Found active BudgetPeriod: {budget_period.name} and Organization: {organization.name}")

            # دریافت تخصیص‌های بودجه
            allocations = BudgetAllocation.objects.filter(
                budget_period=budget_period,
                organization=organization,
                is_active=True
            ).select_related('budget_item').annotate(
                consumed=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='CONSUMPTION'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                ),
                returned=Coalesce(
                    Subquery(
                        BudgetTransaction.objects.filter(
                            allocation_id=OuterRef('pk'),
                            transaction_type='RETURN'
                        ).values('allocation_id').annotate(total=Sum('amount')).values('total')
                    ), Decimal('0'), output_field=DecimalField()
                )
            ).order_by('budget_item__name')

            logger.debug(f"Found {allocations.count()} allocations for this period and organization")

            # بررسی وجود تخصیص‌ها
            if not allocations.exists():
                logger.warning("No active BudgetAllocations found for this period and organization")
                html_content = render_to_string('reports/partials/_no_budget_items_found.html',
                                               {'organization': organization, 'period': budget_period})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            response_data = []
            for alloc in allocations:
                net_consumed = alloc.consumed - alloc.returned
                remaining = alloc.allocated_amount - net_consumed
                # نام تخصیص (بدون پروژه، مگر اینکه تأیید شود)
                alloc_name = alloc.budget_item.name or 'سرفصل نامشخص'

                # تنخواه‌های متصل
                tankhahs = Tankhah.objects.filter(
                    budget_allocation=alloc,
                    is_archived=False
                ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')

                logger.debug(f"Found {tankhahs.count()} tankhahs for BudgetAllocation {alloc.pk}")

                tankhah_data = []
                for t in tankhahs:
                    factors = t.factors.all()
                    factor_data = [
                        {
                            'id': f.pk,
                            'number': f.number,
                            'amount_formatted': f"{f.amount or 0:,.0f}",
                            'status_display': f.get_status_display(),
                            'category_name': f.category.name if f.category else "-",
                            'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime(
                                '%Y/%m/%d') if f.date else "-",
                            'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                            'items_count': f.items.count(),
                            'items': [
                                {
                                    'description': item.description,
                                    'amount_formatted': f"{item.amount or 0:,.0f}",
                                    'quantity': item.quantity,
                                    'unit_price': f"{item.unit_price or 0:,.0f}",
                                    'status_display': item.get_status_display(),
                                } for item in f.items.all()
                            ]
                        } for f in factors
                    ]

                    tankhah_data.append({
                        'id': t.pk,
                        'number': t.number,
                        'amount_formatted': f"{t.amount or 0:,.0f}",
                        'status_display': t.get_status_display(),
                        'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime(
                            '%Y/%m/%d') if t.date else '-',
                        'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                        'factors': factor_data,
                        'factors_count': len(factor_data),
                        'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    })

                response_data.append({
                    'ba_pk': alloc.pk,
                    'budget_item_name': alloc_name,
                    'budget_item_code': alloc.budget_item.code or '-',
                    'ba_allocated_amount_formatted': f"{alloc.allocated_amount or 0:,.0f}",
                    'ba_consumed_amount_formatted': f"{net_consumed or 0:,.0f}",
                    'ba_remaining_amount_formatted': f"{remaining or 0:,.0f}",
                    'ba_utilization_percentage': (net_consumed / alloc.allocated_amount * 100) if alloc.allocated_amount else 0,
                    'tankhahs': tankhah_data,
                    'tankhahs_count': len(tankhah_data),
                    'tankhahs_ajax_url': reverse('api_tankhahs_for_allocation', kwargs={'alloc_pk': alloc.pk}),
                    'add_tankhah_url': reverse('tankhah_create') + f"?budget_allocation_id={alloc.pk}",
                    'ba_report_url': reverse('budget_allocation_report', kwargs={'pk': alloc.pk}),
                })

            logger.info("Successfully prepared budget allocation details")
            html_content = render_to_string(
                'reports/partials/_ajax_level_budget_allocations.html',
                {'budget_allocations_data': response_data, 'parent_period_pk': period_pk, 'parent_org_pk': org_pk}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})

        except Http404 as e:
            logger.error(f"404 Error: {str(e)}")
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': str(e)}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                 {'error_message': _("خطا در پردازش سرفصل‌ها.")}),
                 'status': 'error'},
                status=500
            )
#---------------------------------------------------------------------
class ___APITankhahsForAllocationView(PermissionBaseView, View):
    def get(self, request, alloc_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت تنخواه‌ها برای تخصیص PK={alloc_pk}")
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=alloc_pk, is_active=True)
            tankhahs = Tankhah.objects.filter(
                budget_allocation=allocation,
                is_archived=False
            ).select_related('created_by', 'current_stage').prefetch_related('factors').order_by('-date', '-pk')

            if not tankhahs.exists():
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تخصیص یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                 'status': 'error'},
                status=500
            )
class APITankhahsForAllocationView(View):  # PermissionBaseView حذف شد
    def get(self, request, alloc_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت تنخواه‌ها برای تخصیص PK={alloc_pk}")
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=alloc_pk, is_active=True)
            tankhahs = Tankhah.objects.filter(
                budget_allocation=allocation,
                is_archived=False
            ).select_related('created_by').prefetch_related('factors').order_by('-date', '-pk')

            if not tankhahs.exists():
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تخصیص یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {str(e)}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                 'status': 'error'},
                status=500
            )
#---------------------------------------------------------------------
class APIOrganizationsForPeriodView (PermissionBaseView, View):
    def get(self, request, period_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت سازمان‌ها برای دوره بودجه PK={period_pk}")
        try:
            period = get_object_or_404(BudgetPeriod, pk=period_pk)

            consumed_subquery = BudgetTransaction.objects.filter(
                allocation__organization_id=OuterRef('organization__id'), allocation__budget_period_id=period.pk,
                transaction_type='CONSUMPTION').values('allocation__organization_id').annotate(
                total=Sum('amount')).values('total')
            returned_subquery = BudgetTransaction.objects.filter(
                allocation__organization_id=OuterRef('organization__id'), allocation__budget_period_id=period.pk,
                transaction_type='RETURN').values('allocation__organization_id').annotate(total=Sum('amount')).values(
                'total')

            org_summaries_qs = BudgetAllocation.objects.filter(
                budget_period=period, is_active=True
            ).values(
                'organization__id', 'organization__name', 'organization__code'
            ).annotate(
                total_allocated_to_org=Sum('allocated_amount'),
                num_budget_items_for_org=Count('budget_item', distinct=True),
                total_consumed_for_org=Coalesce(Subquery(consumed_subquery), Decimal('0'), output_field=DecimalField()),
                total_returned_for_org=Coalesce(Subquery(returned_subquery), Decimal('0'), output_field=DecimalField())
            ).order_by('organization__name')

            logger.debug(f"تعداد خلاصه‌های سازمانی یافت شده: {org_summaries_qs.count()}")

            organizations_data = []
            for org_sum in org_summaries_qs:
                net_consumed = org_sum['total_consumed_for_org'] - org_sum['total_returned_for_org']
                remaining = org_sum['total_allocated_to_org'] - net_consumed
                organizations_data.append({
                    'id': org_sum['organization__id'],
                    'name': org_sum['organization__name'],
                    'code': org_sum['organization__code'],
                    'total_allocated_formatted': f"{org_sum['total_allocated_to_org']:,.0f}",
                    'net_consumed_formatted': f"{net_consumed:,.0f}",
                    'remaining_formatted': f"{remaining:,.0f}",
                    'utilization_percentage': (net_consumed / org_sum['total_allocated_to_org'] * 100) if org_sum[
                                                                                                              'total_allocated_to_org'] > 0 else 0,
                    'budget_items_ajax_url': reverse('api_budget_items_for_org_period',
                                                     kwargs={'period_pk': period.pk,
                                                             'org_pk': org_sum['organization__id']})
                })

            logger.info(f"تعداد {len(organizations_data)} سازمان برای رندر در دوره {period_pk} آماده شد.")
            html_content = render_to_string(
                'reports/partials/_ajax_level_organizations.html',
                {'organizations': organizations_data, 'parent_period_pk': period_pk}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            logger.warning(f"API Error: دوره بودجه با PK={period_pk} یافت نشد.")
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("دوره بودجه یافت نشد.")}</p>',
                 'status': 'notfound'}, status=404)
        except Exception as e:
            logger.error(f"خطای بحرانی در APIOrganizationsForPeriodView برای period_pk={period_pk}: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': f'<p class="text-danger p-3 text-center">{_("خطا در پردازش درخواست.")}</p>',
                 'status': 'error'}, status=500)
#----------------------------------------------------------------------
class APITankhahsForPBAView(PermissionBaseView, View):
    def get(self, request, pba_pk, *args, **kwargs):
        logger.info(f"API Request: دریافت تنخواه‌ها برای تخصیص PK={pba_pk}")
        try:
            allocation = get_object_or_404(BudgetAllocation, pk=pba_pk, is_active=True)

            # بررسی اینکه آیا تخصیص پروژه است یا تخصیص اصلی
            if allocation.project:
                tankhahs = Tankhah.objects.filter(
                    project_budget_allocation=allocation,
                    is_archived=False
                )
            else:
                tankhahs = Tankhah.objects.filter(
                    budget_allocation=allocation,
                    is_archived=False
                )

            tankhahs = tankhahs.select_related('created_by', 'current_stage').prefetch_related('factors').order_by(
                '-date', '-pk')

            if not tankhahs.exists():
                return JsonResponse({
                    'html_content': '<p class="text-center p-3">{% translate "هیچ تنخواهی برای این تخصیص یافت نشد." %}</p>',
                    'status': 'empty'
                })

            tankhahs_data = []
            for t in tankhahs:
                tankhahs_data.append({
                    'id': t.pk,
                    'number': t.number,
                    'amount_formatted': f"{t.amount or 0:,.0f}",
                    'status_display': t.get_status_display(),
                    'date_jalali': jdatetime.date.fromgregorian(date=t.date).strftime('%Y/%m/%d') if t.date else "-",
                    'factors_count': t.factors.count(),
                    'factors_ajax_url': reverse('api_factors_for_tankhah', kwargs={'tankhah_pk': t.pk}),
                    'detail_url': reverse('tankhah_detail', kwargs={'pk': t.pk}),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_tankhahs.html',
                {'tankhahs_list_data': tankhahs_data}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html',
                                                  {'message': _("تخصیص یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html',
                                                  {'error_message': _("خطا در پردازش تنخواه‌ها.")}),
                 'status': 'error'},
                status=500
            )
#---------------------------------------------------------------------
class APIFactorsForTankhahView(PermissionBaseView, View):
    def get(self, request, tankhah_pk, *args, **kwargs):
        logger.info(f"API request: Get Factors for Tankhah PK={tankhah_pk}")
        try:
            tankhah = get_object_or_404(Tankhah.objects.select_related('organization'), pk=tankhah_pk)
            factors = Factor.objects.filter(tankhah=tankhah).select_related('category', 'created_by').prefetch_related('items').order_by('-date')

            if not factors.exists():
                html_content = render_to_string('reports/partials/_no_factors_found.html', {'parent_tankhah': tankhah})
                return JsonResponse({'html_content': html_content, 'status': 'empty'})

            factors_data = []
            for f in factors:
                items = f.items.all()
                item_data = [
                    {
                        'description': item.description,
                        'amount_formatted': f"{item.amount or 0:,.0f}",
                        'quantity': item.quantity,
                        'unit_price': f"{item.unit_price or 0:,.0f}",
                        'status_display': item.get_status_display(),
                    } for item in items
                ]

                factors_data.append({
                    'id': f.pk,
                    'number': f.number,
                    'amount_formatted': f"{f.amount or 0:,.0f}",
                    'status_display': f.get_status_display(),
                    'category_name': f.category.name if f.category else "-",
                    'date_jalali': jdatetime.date.fromgregorian(date=f.date).strftime('%Y/%m/%d') if f.date else "-",
                    'detail_url': reverse('factor_detail', kwargs={'factor_pk': f.pk}),
                    'items': item_data,
                    'items_count': len(item_data),
                })

            html_content = render_to_string(
                'reports/partials/_ajax_level_factors.html',
                {'factors_list_data': factors_data, 'parent_tankhah_pk': tankhah_pk, 'parent_tankhah_number': tankhah.number}
            )
            return JsonResponse({'html_content': html_content, 'status': 'success'})
        except Http404:
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_no_data_found_ajax.html', {'message': _("تنخواه یافت نشد.")}),
                 'status': 'notfound'},
                status=404
            )
        except Exception as e:
            logger.error(f"API Error: {e}", exc_info=True)
            return JsonResponse(
                {'html_content': render_to_string('reports/partials/_error_ajax.html', {'error_message': _("خطا در پردازش فاکتورها.")}),
                 'status': 'error'},
                status=500
            )

# ==============================================================================
# ویوی اصلی گزارش جامع (سطح صفر: دوره‌های بودجه)
# ==============================================================================

# ==============================================================================
# API سطح اول: دریافت سازمان‌های یک دوره بودجه
# ==============================================================================
# ==============================================================================
# سایر ویوهای API (با الگوی لاگ مشابه)
# ==============================================================================
